from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import openpyxl

from crs_fatca_generator.services.controlling_person_service import detect_controlling_person_blocks


@dataclass
class ExcelPreview:
    path: Path
    sheets: list[str]
    active_sheet: str
    headers: list[str]
    rows: list[dict[str, Any]]
    duplicate_headers: list[str]
    column_types: dict[str, str]


class ExcelReader:
    def list_sheets(self, path: Path) -> list[str]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        try:
            return wb.sheetnames
        finally:
            wb.close()

    def preview(
        self,
        path: Path,
        sheet_name: str | None = None,
        header_row: int = 1,
        start_row: int | None = None,
        end_row: int | None = None,
        limit: int = 20,
    ) -> ExcelPreview:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        try:
            sheet = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
            raw_headers = self._headers(sheet, header_row)
            headers = self._unique_headers(raw_headers)
            duplicate_headers = self._duplicates(raw_headers)
            needs_raw_values = _needs_raw_values(raw_headers)
            first_data_row = start_row or header_row + 1
            last_row = min(end_row or sheet.max_row, first_data_row + limit - 1)
            rows = []
            type_samples: dict[str, list[str]] = {header: [] for header in headers}
            for row in sheet.iter_rows(min_row=first_data_row, max_row=last_row, values_only=True):
                item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
                item["_headers"] = raw_headers
                item["_field_headers"] = headers
                if needs_raw_values:
                    item["_raw_values"] = row[: len(headers)]
                rows.append(item)
                for header, value in item.items():
                    if header.startswith("_"):
                        continue
                    if value is not None:
                        type_samples[header].append(type(value).__name__)
            column_types = {
                header: (max(set(samples), key=samples.count) if samples else "vazio")
                for header, samples in type_samples.items()
            }
            return ExcelPreview(path, wb.sheetnames, sheet.title, headers, rows, duplicate_headers, column_types)
        finally:
            wb.close()

    def read_rows(
        self,
        path: Path,
        sheet_name: str,
        header_row: int,
        start_row: int | None = None,
        end_row: int | None = None,
        progress_callback: Callable[[int, int, int, dict[str, Any]], None] | None = None,
    ) -> list[dict[str, Any]]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        try:
            sheet = wb[sheet_name]
            raw_headers = self._headers(sheet, header_row)
            headers = self._unique_headers(raw_headers)
            needs_raw_values = _needs_raw_values(raw_headers)
            first_data_row = start_row or header_row + 1
            last_row = end_row or sheet.max_row
            total_rows = max(last_row - first_data_row + 1, 0)
            rows: list[dict[str, Any]] = []
            for index, row in enumerate(sheet.iter_rows(min_row=first_data_row, max_row=last_row, values_only=True), first_data_row):
                item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
                item["_excel_row"] = index
                item["_headers"] = raw_headers
                item["_field_headers"] = headers
                if needs_raw_values:
                    item["_raw_values"] = row[: len(headers)]
                if any(value not in (None, "") for key, value in item.items() if not key.startswith("_")):
                    rows.append(item)
                processed = index - first_data_row + 1
                if progress_callback and (processed == 1 or processed % 100 == 0 or processed == total_rows):
                    progress_callback(processed, total_rows, index, item)
            return rows
        finally:
            wb.close()

    def _headers(self, sheet: Any, header_row: int) -> list[str]:
        raw = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        headers: list[str] = []
        for idx, value in enumerate(raw, 1):
            name = str(value).strip() if value not in (None, "") else f"Coluna {idx}"
            headers.append(name)
        while headers and headers[-1].startswith("Coluna "):
            headers.pop()
        return headers

    def _duplicates(self, headers: list[str]) -> list[str]:
        seen: set[str] = set()
        dupes: set[str] = set()
        for header in headers:
            key = header.casefold()
            if key in seen:
                dupes.add(header)
            seen.add(key)
        return sorted(dupes)

    def _unique_headers(self, headers: list[str]) -> list[str]:
        counts: dict[str, int] = {}
        unique: list[str] = []
        for header in headers:
            key = header.casefold()
            counts[key] = counts.get(key, 0) + 1
            unique.append(header if counts[key] == 1 else f"{header} #{counts[key]}")
        return unique


def _needs_raw_values(raw_headers: list[str]) -> bool:
    return bool(detect_controlling_person_blocks(raw_headers))
