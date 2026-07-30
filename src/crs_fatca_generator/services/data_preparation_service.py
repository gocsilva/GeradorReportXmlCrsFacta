from __future__ import annotations

import csv
import json
import logging
import os
import platform
from uuid import uuid4
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

import openpyxl
from openpyxl.utils import get_column_letter

from crs_fatca_generator.models.domain import ValidationIssue
from crs_fatca_generator.models.mapping import MappingProfile
from crs_fatca_generator.security.masking import mask_value
from crs_fatca_generator.services.controlling_person_service import ControllingPersonRecord, extract_controlling_persons
from crs_fatca_generator.services.fatca_missing_tin_policy import FatcaMissingTinPolicy
from crs_fatca_generator.services.file_hash import sha256_file
from crs_fatca_generator.services.tax_identifier_service import classify_tax_identifier
from crs_fatca_generator.services.transformation_service import to_date
from crs_fatca_generator.services.transformation_service import is_empty, normalize_text


CLOSED_CC_STATUS = {"encerrada", "encerrada bacen"}
PREPARATION_PROGRESS_INTERVAL = 1000
AUDIT_FULL_XLSX_ROW_LIMIT = 50_000
AUDIT_XLSX_SAMPLE_LIMIT = 500

ProgressCallback = Callable[[dict[str, Any]], None]
logger = logging.getLogger(__name__)


@dataclass
class AuditEvent:
    event_type: str
    rule: str
    excel_row: int | None
    document: str = ""
    account: str = ""
    person_type: str = ""
    action: str = ""
    detail: str = ""
    field: str = ""
    original_value: str = ""
    intermediate_value: str = ""
    final_value: str = ""
    transformation: str = ""
    result: str = ""
    severity: str = "INFO"
    evidence: str = ""
    required_data: str = ""
    found_data: str = ""

    def as_row(self) -> dict[str, object]:
        return {
            "tipo": self.event_type,
            "regra": self.rule,
            "linha_excel": self.excel_row or "",
            "documento_mascarado": mask_value(self.document),
            "conta_mascarada": mask_value(self.account),
            "tipo_pessoa": self.person_type,
            "acao": self.action,
            "detalhe": self.detail,
            "campo": self.field,
            "valor_original": self.original_value,
            "valor_intermediario": self.intermediate_value,
            "valor_final": self.final_value,
            "transformacao": self.transformation,
            "resultado": self.result,
            "severidade": self.severity,
            "evidencia": self.evidence,
        }


@dataclass
class PreparedData:
    rows: list[dict[str, Any]]
    original_rows: list[dict[str, Any]] = field(default_factory=list)
    events: list[AuditEvent] = field(default_factory=list)
    issues: list[ValidationIssue] = field(default_factory=list)
    processing_id: str = field(default_factory=lambda: uuid4().hex)


class DataPreparationService:
    def prepare(
        self,
        rows: list[dict[str, Any]],
        profile: MappingProfile,
        progress_callback: ProgressCallback | None = None,
    ) -> PreparedData:
        working = [dict(row) for row in rows]
        original_rows = [dict(row) for row in rows]
        events: list[AuditEvent] = []
        period_start = _reporting_start(profile)
        period_year = period_start.year

        self._emit_progress(progress_callback, "Preparando dados: regras de encerramento CI", 0, len(working), "")
        working = self._remove_closed_before_period(working, events, period_start, progress_callback)
        self._emit_progress(progress_callback, "Preparando dados: regras de encerramento CC", 0, len(working), "")
        working = self._remove_cc_closed_during_period(working, events, period_year, progress_callback)
        self._emit_progress(progress_callback, "Preparando dados: documentos", 0, len(working), "")
        issues = self._normalize_documents(working, profile, events, progress_callback)
        self._emit_progress(progress_callback, "Preparando dados: controladores", 0, len(working), "")
        issues.extend(self._prepare_controlling_persons(working, events, progress_callback))
        if not issues:
            working = self._finalize_valid_rows(working, profile, events, progress_callback)
        self._emit_progress(progress_callback, "Preparando dados: finalizado", len(working), len(working), "")
        return PreparedData(working, original_rows, events, issues)

    def ignore_issue_rows(
        self,
        prepared: PreparedData,
        profile: MappingProfile,
        progress_callback: ProgressCallback | None = None,
    ) -> PreparedData:
        issue_lines = {issue.excel_row for issue in prepared.issues if issue.level == "erro" and issue.excel_row}
        remaining_issues = [issue for issue in prepared.issues if not issue.excel_row or issue.excel_row not in issue_lines]
        if not issue_lines:
            return prepared
        source_by_line = self._prepared_by_line(prepared.original_rows)
        events = list(prepared.events)
        for line in sorted(issue_lines):
            source = source_by_line.get(line)
            details = "; ".join(issue.message for issue in prepared.issues if issue.excel_row == line)
            if source:
                events.append(
                    self._event(
                        source,
                        "REMOVIDO",
                        "ERRO_IGNORADO",
                        "registro_ignorado",
                        f"Registro removido da geracao por erro nao corrigivel: {details}",
                        result="IGNORADO_PELO_USUARIO",
                        severity="BLOQUEIO",
                        evidence=f"linha_excel={line}",
                    )
                )
        kept_rows = [row for row in prepared.rows if _excel_row(row) not in issue_lines]
        self._emit_progress(
            progress_callback,
            "Preparando dados: ignorando registros com erro",
            len(issue_lines),
            len(issue_lines),
            f"{len(issue_lines)} registros ignorados",
        )
        if not remaining_issues:
            kept_rows = self._finalize_valid_rows(kept_rows, profile, events, progress_callback)
        return PreparedData(kept_rows, prepared.original_rows, events, remaining_issues, prepared.processing_id)

    def _finalize_valid_rows(
        self,
        working: list[dict[str, Any]],
        profile: MappingProfile,
        events: list[AuditEvent],
        progress_callback: ProgressCallback | None,
    ) -> list[dict[str, Any]]:
        self._emit_progress(progress_callback, "Preparando dados: contas duplicadas", 0, len(working), "")
        working = self._remove_lowest_duplicate_account(working, profile, events, progress_callback)
        self._emit_progress(progress_callback, "Preparando dados: saldos negativos", 0, len(working), "")
        self._zero_negative_balances(working, profile, events, progress_callback)
        self._emit_progress(progress_callback, "Preparando dados: FATCA US Tax ID", 0, len(working), "")
        self._audit_fatca_us_tin(working, profile, events, progress_callback)
        return working

    def write_audit(
        self,
        prepared: PreparedData,
        output_dir: Path,
        stem: str,
        profile: MappingProfile | None = None,
        excel_path: Path | None = None,
        file_hash: str = "",
        results: list[Any] | None = None,
        reports: dict[str, Any] | None = None,
        progress_callback: ProgressCallback | None = None,
    ) -> tuple[Path, Path, Path]:
        output_dir.mkdir(parents=True, exist_ok=True)
        csv_path = output_dir / f"{stem}_relatorio_auditoria.csv"
        xlsx_path = output_dir / f"{stem}_relatorio_auditoria.xlsx"
        json_path = output_dir / f"{stem}_manifesto_auditoria.json"
        reports = reports or {}
        results = results or []
        rows = self._csv_rows(prepared, results, reports)
        headers = list(rows[0].keys())
        self._emit_progress(progress_callback, "Gerando auditoria: CSV completo", 0, len(rows), csv_path.name)
        with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers, delimiter=";")
            writer.writeheader()
            for index, row in enumerate(rows, 1):
                writer.writerow(row)
                self._emit_progress(progress_callback, "Gerando auditoria: CSV completo", index, len(rows), f"CSV linha {index}", force=index == len(rows))
        workbook = openpyxl.Workbook()
        compact_audit = self._use_compact_xlsx(prepared)
        if compact_audit:
            self._write_compact_xlsx(workbook, prepared, profile, excel_path, file_hash, results, reports, csv_path, xlsx_path, json_path, progress_callback)
        else:
            self._replace_sheet(workbook, "Resumo", self._summary_rows(prepared, profile, excel_path, file_hash, results, reports), progress_callback)
            self._replace_sheet(workbook, "Entrada", self._input_rows(prepared), progress_callback)
            self._replace_sheet(workbook, "Decisoes", self._decision_rows(prepared), progress_callback)
            self._replace_sheet(workbook, "Avaliacao_Regras", self._rule_evaluation_rows(prepared), progress_callback)
            self._replace_sheet(workbook, "Exclusoes", self._exclusion_rows(prepared), progress_callback)
            self._replace_sheet(workbook, "Transformacoes", self._transformation_rows(prepared), progress_callback)
            self._replace_sheet(workbook, "CRS", self._report_rows(reports.get("crs"), "crs"), progress_callback)
            self._replace_sheet(workbook, "FATCA", self._report_rows(reports.get("fatca"), "fatca"), progress_callback)
            self._replace_sheet(workbook, "ControllingPersons", self._controlling_person_rows(prepared, reports), progress_callback)
            self._replace_sheet(workbook, "Identificadores", self._identifier_rows(reports), progress_callback)
            self._replace_sheet(workbook, "Validacao_XSD", self._xsd_rows(results), progress_callback)
            self._replace_sheet(workbook, "Conciliacao", self._conciliation_rows(prepared, reports), progress_callback)
            self._replace_sheet(workbook, "Conciliacao_Controladores", self._controlling_conciliation_rows(prepared, reports), progress_callback)
            self._replace_sheet(workbook, "Pendencias", self._pending_rows(prepared), progress_callback)
            self._replace_sheet(workbook, "Hashes", self._hash_rows(prepared, excel_path, file_hash, results, csv_path, xlsx_path, json_path), progress_callback)
            self._replace_sheet(workbook, "Auditoria", [event.as_row() for event in prepared.events] + self._issue_rows(prepared.issues), progress_callback)
            self._replace_sheet(workbook, "Pendencias US Tax ID", [self._fatca_pending_row(event) for event in prepared.events if event.event_type == "PENDENCIA_US_TAX_ID"], progress_callback)
        self._emit_progress(progress_callback, "Gerando auditoria: salvando XLSX", 0, 1, xlsx_path.name)
        workbook.save(xlsx_path)
        self._emit_progress(progress_callback, "Gerando auditoria: salvando XLSX", 1, 1, xlsx_path.name)
        manifest = self._manifest(prepared, profile, excel_path, file_hash, results, reports, compact=compact_audit)
        if compact_audit:
            manifest["audit_xlsx_mode"] = "resumido"
            manifest["audit_xlsx_reason"] = f"Arquivo grande: XLSX detalhado desativado acima de {AUDIT_FULL_XLSX_ROW_LIMIT} registros/eventos. CSV contem a auditoria completa."
        self._emit_progress(progress_callback, "Gerando auditoria: manifesto", 0, 1, json_path.name)
        json_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
        self._emit_progress(progress_callback, "Gerando auditoria: manifesto", 1, 1, json_path.name)
        return csv_path, xlsx_path, json_path

    def _use_compact_xlsx(self, prepared: PreparedData) -> bool:
        return max(len(prepared.original_rows), len(prepared.rows), len(prepared.events)) > AUDIT_FULL_XLSX_ROW_LIMIT

    def _write_compact_xlsx(
        self,
        workbook: openpyxl.Workbook,
        prepared: PreparedData,
        profile: MappingProfile | None,
        excel_path: Path | None,
        file_hash: str,
        results: list[Any],
        reports: dict[str, Any],
        csv_path: Path,
        xlsx_path: Path,
        json_path: Path,
        progress_callback: ProgressCallback | None,
    ) -> None:
        logger.info(
            "AUDIT_COMPACT_XLSX rows_original=%s rows_prepared=%s events=%s csv=%s",
            len(prepared.original_rows),
            len(prepared.rows),
            len(prepared.events),
            csv_path,
        )
        sample = self._sample_prepared(prepared, AUDIT_XLSX_SAMPLE_LIMIT)
        summary = self._summary_rows(prepared, profile, excel_path, file_hash, results, reports)
        self._replace_sheet(workbook, "Resumo", summary, progress_callback)
        self._replace_sheet(
            workbook,
            "Modo_XLSX",
            [
                {
                    "modo": "resumido",
                    "limite_xlsx_completo": AUDIT_FULL_XLSX_ROW_LIMIT,
                    "limite_amostra": AUDIT_XLSX_SAMPLE_LIMIT,
                    "registros_recebidos": len(prepared.original_rows),
                    "registros_preparados": len(prepared.rows),
                    "eventos_auditoria": len(prepared.events),
                    "observacao": f"Auditoria completa linha a linha disponivel no CSV: {csv_path}",
                }
            ],
            progress_callback,
        )
        self._replace_sheet(
            workbook,
            "Arquivos",
            [
                {"tipo": "CSV completo", "caminho": str(csv_path), "observacao": "Auditoria completa linha a linha."},
                {"tipo": "XLSX resumido", "caminho": str(xlsx_path), "observacao": f"Amostras limitadas a {AUDIT_XLSX_SAMPLE_LIMIT} linhas por aba."},
                {"tipo": "Manifesto JSON", "caminho": str(json_path), "observacao": "Resumo tecnico e hashes."},
            ],
            progress_callback,
        )
        self._replace_sheet(workbook, "Entrada_Amostra", self._input_rows(sample), progress_callback)
        self._replace_sheet(workbook, "Decisoes_Amostra", self._decision_rows(sample), progress_callback)
        self._replace_sheet(workbook, "Auditoria_Amostra", [event.as_row() for event in sample.events] + self._issue_rows(sample.issues), progress_callback)
        self._replace_sheet(workbook, "Identificadores_Resumo", self._identifier_summary_rows(reports), progress_callback)
        self._replace_sheet(workbook, "Validacao_XSD", self._xsd_rows(results), progress_callback)

    def _sample_prepared(self, prepared: PreparedData, limit: int) -> PreparedData:
        return PreparedData(
            rows=prepared.rows[:limit],
            original_rows=prepared.original_rows[:limit],
            events=prepared.events[:limit],
            issues=prepared.issues[:limit],
            processing_id=prepared.processing_id,
        )

    def _remove_closed_before_period(
        self,
        rows: list[dict[str, Any]],
        events: list[AuditEvent],
        period_start: date,
        progress_callback: ProgressCallback | None = None,
    ) -> list[dict[str, Any]]:
        kept: list[dict[str, Any]] = []
        has_ci_column = _has_column(rows, "DataHoraEncerramento CI")
        if not has_ci_column:
            events.append(
                self._summary_rule_event(
                    "REGRA_01",
                    "NAO_AVALIADA_DADO_AUSENTE",
                    f"Coluna ausente: DataHoraEncerramento CI. {len(rows)} registros mantidos sem avaliacao individual.",
                    "ALERTA",
                    required_data="DataHoraEncerramento CI",
                    found_data="nao",
                )
            )
            self._emit_progress(progress_callback, "Preparando dados: regras de encerramento CI", len(rows), len(rows), "")
            return rows
        total = len(rows)
        for index, row in enumerate(rows, 1):
            closed_at = _parse_date(row.get("DataHoraEncerramento CI"))
            if closed_at and closed_at < period_start:
                events.append(self._rule_event(row, "REGRA_01", "APLICADA_COM_EXCLUSAO", f"Encerramento CI em {closed_at.isoformat()} antes de {period_start.isoformat()}.", "BLOQUEIO", original_value=normalize_text(row.get("DataHoraEncerramento CI")), final_value=closed_at.isoformat(), evidence=f"{closed_at.isoformat()} < {period_start.isoformat()}", required_data="DataHoraEncerramento CI", found_data="sim"))
                events.append(self._event(row, "REMOVIDO", "REGRA_01", "conta_removida", f"Encerramento CI em {closed_at.isoformat()} antes de {period_start.isoformat()}."))
                self._emit_progress(progress_callback, "Preparando dados: regras de encerramento CI", index, total, self._row_label(row), force=index == total)
                continue
            result = "AVALIADA_NAO_APLICAVEL"
            reason = "Sem encerramento CI anterior ao inicio do periodo."
            events.append(self._rule_event(row, "REGRA_01", result, reason, "INFO", original_value=normalize_text(row.get("DataHoraEncerramento CI")), final_value=closed_at.isoformat() if closed_at else "", evidence=f"inicio_periodo={period_start.isoformat()}", required_data="DataHoraEncerramento CI", found_data="sim"))
            kept.append(row)
            self._emit_progress(progress_callback, "Preparando dados: regras de encerramento CI", index, total, self._row_label(row), force=index == total)
        return kept

    def _remove_cc_closed_during_period(
        self,
        rows: list[dict[str, Any]],
        events: list[AuditEvent],
        period_year: int,
        progress_callback: ProgressCallback | None = None,
    ) -> list[dict[str, Any]]:
        kept: list[dict[str, Any]] = []
        has_required = _has_column(rows, "DataHoraEncerramento CI") and _has_column(rows, "Encerramento CC") and _has_column(rows, "Status em 31/12 em CC")
        if not has_required:
            missing = [name for name in ("DataHoraEncerramento CI", "Encerramento CC", "Status em 31/12 em CC") if not _has_column(rows, name)]
            events.append(
                self._summary_rule_event(
                    "REGRA_03",
                    "NAO_AVALIADA_DADO_AUSENTE",
                    "Colunas ausentes: " + ", ".join(missing) + f". {len(rows)} registros mantidos sem avaliacao individual.",
                    "ALERTA",
                    required_data="DataHoraEncerramento CI;Encerramento CC;Status em 31/12 em CC",
                    found_data="nao",
                )
            )
            self._emit_progress(progress_callback, "Preparando dados: regras de encerramento CC", len(rows), len(rows), "")
            return rows
        total = len(rows)
        for index, row in enumerate(rows, 1):
            ci_closed_at = _parse_date(row.get("DataHoraEncerramento CI"))
            cc_closed_at = _parse_date(row.get("Encerramento CC"))
            status = _normalize_status(row.get("Status em 31/12 em CC"))
            if not ci_closed_at and status in CLOSED_CC_STATUS and cc_closed_at and cc_closed_at.year == period_year:
                detail = f"Status CC {status} em {cc_closed_at.isoformat()} sem encerramento efetivo no CI."
                events.append(self._rule_event(row, "REGRA_03", "APLICADA_COM_EXCLUSAO", detail, "BLOQUEIO", original_value=normalize_text(row.get("Status em 31/12 em CC")), final_value=status, evidence=f"ano_status={cc_closed_at.year}; ano_periodo={period_year}", required_data="DataHoraEncerramento CI;Encerramento CC;Status em 31/12 em CC", found_data="sim"))
                events.append(self._event(row, "REMOVIDO", "REGRA_03", "conta_removida", detail))
                self._emit_progress(progress_callback, "Preparando dados: regras de encerramento CC", index, total, self._row_label(row), force=index == total)
                continue
            events.append(self._rule_event(row, "REGRA_03", "AVALIADA_NAO_APLICAVEL", "Status CC/CI nao atende criterio de exclusao.", "INFO", original_value=normalize_text(row.get("Status em 31/12 em CC")), final_value=status, evidence=f"ci={ci_closed_at}; cc={cc_closed_at}; ano_periodo={period_year}", required_data="DataHoraEncerramento CI;Encerramento CC;Status em 31/12 em CC", found_data="sim"))
            kept.append(row)
            self._emit_progress(progress_callback, "Preparando dados: regras de encerramento CC", index, total, self._row_label(row), force=index == total)
        return kept

    def _normalize_documents(
        self,
        rows: list[dict[str, Any]],
        profile: MappingProfile,
        events: list[AuditEvent],
        progress_callback: ProgressCallback | None = None,
    ) -> list[ValidationIssue]:
        document_column = _column_for(profile, "holder.tin", "DocumentoCliente")
        person_column = _column_for(profile, "holder.kind", "Tipo de documento")
        country_column = _column_for(profile, "holder.res_country", "Pais")
        issues: list[ValidationIssue] = []
        if not _has_column(rows, document_column) or not _has_column(rows, person_column):
            return issues
        total = len(rows)
        for index, row in enumerate(rows, 1):
            try:
                classified = classify_tax_identifier(row.get(document_column), row.get(person_column), row.get(country_column, "BR"))
            except ValueError as exc:
                issues.append(ValidationIssue("erro", "DOCSCI001", str(exc), document_column, _excel_row(row), "Corrija o documento no Excel e reexecute."))
                self._emit_progress(progress_callback, "Preparando dados: documentos", index, total, self._row_label(row), force=index == total)
                continue
            if not classified.valid:
                issues.append(ValidationIssue("erro", "DOC001", classified.message, document_column, _excel_row(row), "Corrija CPF/CNPJ no Excel."))
                self._emit_progress(progress_callback, "Preparando dados: documentos", index, total, self._row_label(row), force=index == total)
                continue
            raw_document = normalize_text(row.get(document_column))
            digits = "".join(char for char in raw_document if char.isdigit())
            if normalize_text(row.get(document_column)) != classified.normalized:
                added = max(0, len(classified.normalized) - len(digits))
                events.append(
                    self._event(
                        row,
                        "AJUSTE",
                        "DOC_NORMALIZACAO",
                        "documento_normalizado",
                        f"{classified.kind} normalizado para {mask_value(classified.normalized)}.",
                        field=classified.kind.lower(),
                        original_value=raw_document,
                        intermediate_value=digits,
                        final_value=classified.normalized,
                        transformation="ZFILL" if added else "STRIP_NON_DIGITS",
                        result="VALIDO",
                        severity="INFO",
                        evidence=f"zeros_adicionados={added}; validacao_matematica=valida",
                    )
                )
                row[document_column] = classified.normalized
            row["_tax_identifier_kind"] = classified.kind
            row["_tax_identifier_issued_by"] = classified.issued_by
            row["_documento_brasileiro"] = classified.normalized if classified.issued_by == "BR" else ""
            row["_tipo_documento_brasileiro"] = classified.kind if classified.issued_by == "BR" else ""
            row["_cpf"] = classified.normalized if classified.kind == "CPF" else ""
            row["_cnpj"] = classified.normalized if classified.kind == "CNPJ" else ""
            self._emit_progress(progress_callback, "Preparando dados: documentos", index, total, self._row_label(row), force=index == total)
        return issues

    def _prepare_controlling_persons(
        self,
        rows: list[dict[str, Any]],
        events: list[AuditEvent],
        progress_callback: ProgressCallback | None = None,
    ) -> list[ValidationIssue]:
        issues: list[ValidationIssue] = []
        total = len(rows)
        for index, row in enumerate(rows, 1):
            records, record_issues, metrics = extract_controlling_persons(row)
            included = [record for record in records if record.result == "INCLUIDO"]
            row["_controlling_persons"] = included
            row["_controlling_person_audit"] = records
            row["_controlling_person_metrics"] = metrics
            for record in records:
                events.append(
                    AuditEvent(
                        event_type="CONTROLLING_PERSON",
                        rule="CONTROLLING_PERSON",
                        excel_row=record.excel_row,
                        document=record.holder_document,
                        account=record.account_number,
                        person_type="PJ",
                        action="incluido" if record.result == "INCLUIDO" else "excluido",
                        detail=record.reason or "Controlador valido para CRS.",
                        field=f"ControllingPerson bloco {record.block_index}",
                        original_value=record.raw_tin,
                        intermediate_value="".join(char for char in record.raw_tin if char.isdigit()),
                        final_value=record.normalized_tin,
                        transformation="ZFILL" if record.normalized else "NONE",
                        result=record.result,
                        severity="INFO" if record.result == "INCLUIDO" else "BLOQUEIO",
                        evidence=f"tipo={record.controlling_person_type}; pais={record.tax_residence}; issuedBy={record.tin_issued_by}",
                    )
                )
            issues.extend(record_issues)
            self._emit_progress(progress_callback, "Preparando dados: controladores", index, total, self._row_label(row), force=index == total)
        return issues

    def _remove_lowest_duplicate_account(
        self,
        rows: list[dict[str, Any]],
        profile: MappingProfile,
        events: list[AuditEvent],
        progress_callback: ProgressCallback | None = None,
    ) -> list[dict[str, Any]]:
        document_column = _column_for(profile, "holder.tin", "DocumentoCliente")
        account_column = _column_for(profile, "account.account_number", "NumConta")
        groups: dict[str, list[dict[str, Any]]] = {}
        total_rows = len(rows)
        for index, row in enumerate(rows, 1):
            document = normalize_text(row.get(document_column))
            if document:
                groups.setdefault(document, []).append(row)
            self._emit_progress(progress_callback, "Preparando dados: contas duplicadas", index, total_rows, self._row_label(row), force=False)
        remove_ids: set[int] = set()
        single_account_groups = 0
        total_groups = len(groups)
        for index, group_rows in enumerate(groups.values(), 1):
            if len(group_rows) <= 1:
                row = group_rows[0]
                single_account_groups += 1
                self._emit_progress(progress_callback, "Preparando dados: contas duplicadas", index, total_groups, self._row_label(row), force=index == total_groups)
                continue
            sorted_rows = sorted(group_rows, key=lambda item: _account_sort_key(item.get(account_column)))
            lowest = sorted_rows[0]
            kept = sorted_rows[-1]
            all_accounts = ",".join(normalize_text(item.get(account_column)) for item in group_rows)
            remove_ids.add(id(lowest))
            for item in group_rows:
                result = "APLICADA_COM_EXCLUSAO" if item is lowest else "AVALIADA_NAO_APLICAVEL"
                detail = "Mesmo documento com mais de uma conta CI; removida a conta de menor numero." if item is lowest else "Mesmo documento com mais de uma conta CI; conta mantida por nao ser a menor."
                events.append(self._rule_event(item, "REGRA_02", result, detail, "BLOQUEIO" if item is lowest else "INFO", original_value=all_accounts, final_value=normalize_text(kept.get(account_column)), evidence=f"contas={all_accounts}; removida={normalize_text(lowest.get(account_column))}; mantida={normalize_text(kept.get(account_column))}; criterio=numerico_quando_possivel", required_data=f"{document_column};{account_column}", found_data="sim"))
            events.append(self._event(lowest, "REMOVIDO", "REGRA_02", "conta_removida", "Mesmo documento com mais de uma conta CI; removida a conta de menor numero."))
            self._emit_progress(progress_callback, "Preparando dados: contas duplicadas", index, total_groups, self._row_label(kept), force=index == total_groups)
        if single_account_groups:
            events.append(
                self._summary_rule_event(
                    "REGRA_02",
                    "AVALIADA_NAO_APLICAVEL",
                    f"{single_account_groups} documentos possuem apenas uma conta no agrupamento avaliado.",
                    "INFO",
                    required_data=f"{document_column};{account_column}",
                    found_data="sim",
                )
            )
        return [row for row in rows if id(row) not in remove_ids]

    def _zero_negative_balances(
        self,
        rows: list[dict[str, Any]],
        profile: MappingProfile,
        events: list[AuditEvent],
        progress_callback: ProgressCallback | None = None,
    ) -> None:
        balance_column = _column_for(profile, "account.balance", "SaldoTotal")
        if not _has_column(rows, balance_column):
            return
        total = len(rows)
        for index, row in enumerate(rows, 1):
            balance = _decimal(row.get(balance_column))
            if balance is not None and balance < 0:
                events.append(
                    self._event(
                        row,
                        "AJUSTE",
                        "SALDO_NEGATIVO",
                        "saldo_zerado",
                        "Saldo negativo convertido para 0.00.",
                        field="saldo",
                        original_value=normalize_text(row.get(balance_column)),
                        intermediate_value=str(balance),
                        final_value="0.00",
                        transformation="NEGATIVE_TO_ZERO",
                        result="AJUSTADO",
                        severity="INFO",
                        evidence="saldo_final=max(saldo_original, Decimal('0.00')); moeda_final=USD; casas_decimais=2",
                    )
                )
                row[balance_column] = "0.00"
            self._emit_progress(progress_callback, "Preparando dados: saldos negativos", index, total, self._row_label(row), force=index == total)

    def _audit_fatca_us_tin(
        self,
        rows: list[dict[str, Any]],
        profile: MappingProfile,
        events: list[AuditEvent],
        progress_callback: ProgressCallback | None = None,
    ) -> None:
        policy = FatcaMissingTinPolicy()
        total = len(rows)
        for index, row in enumerate(rows, 1):
            decision = policy.decide_xml_representation(row, profile)
            row["_fatca_us_tin"] = decision.tins[0].value if decision.tins else ""
            row["_fatca_us_tin_status"] = decision.status
            row["_fatca_us_tin_reason"] = decision.reason
            row["_fatca_us_tin_issued_by"] = decision.tins[0].issued_by if decision.tins else "US"
            row["_fatca_us_tin_policy"] = decision.policy
            row["_fatca_us_tin_blocking"] = "sim" if decision.blocking else "nao"
            if not decision.tins or decision.blocking:
                events.append(
                    AuditEvent(
                        event_type="PENDENCIA_US_TAX_ID",
                        rule="FATCA_US_TIN",
                        excel_row=_excel_row(row),
                        document=normalize_text(row.get("_documento_brasileiro") or _row_value(row, "DocumentoCliente", "Identification Number / CPF")),
                        account=normalize_text(_row_value(row, "NumConta", "AccountNumber")),
                        person_type=normalize_text(_row_value(row, "Tipo de documento", "DocumentType")),
                        action="fatca_us_tin_pendente",
                        detail=decision.treatment,
                        field="us_tin",
                        original_value="",
                        intermediate_value="",
                        final_value="",
                        transformation="OMIT_ELEMENT" if not decision.blocking else "BLOCK_PRODUCTION",
                        result="ALERTA" if not decision.blocking else "BLOQUEIO",
                        severity="ALERTA" if not decision.blocking else "BLOQUEIO",
                        evidence=decision.reason,
                    )
                )
            self._emit_progress(progress_callback, "Preparando dados: FATCA US Tax ID", index, total, self._row_label(row), force=index == total)

    def _emit_progress(
        self,
        progress_callback: ProgressCallback | None,
        phase: str,
        processed: int,
        total: int,
        current_record: str,
        force: bool = True,
    ) -> None:
        if not progress_callback:
            return
        if not force and processed not in {0, 1, total} and processed % PREPARATION_PROGRESS_INTERVAL != 0:
            return
        logger.info("PROGRESS phase=%s processed=%s total=%s current=%s", phase, processed, total, current_record)
        progress_callback(
            {
                "phase": phase,
                "kind": "",
                "processed": processed,
                "total": total,
                "current_record": current_record,
            }
        )

    def _row_label(self, row: dict[str, Any]) -> str:
        excel_row = _excel_row(row)
        account = normalize_text(_row_value(row, "NumConta", "AccountNumber"))
        document = normalize_text(_row_value(row, "DocumentoCliente", "Identification Number / CPF", "_documento_brasileiro"))
        if account:
            return f"linha {excel_row} | conta {account}"
        if document:
            return f"linha {excel_row} | documento {document}"
        return f"linha {excel_row}".strip()

    def _event(
        self,
        row: dict[str, Any],
        event_type: str,
        rule: str,
        action: str,
        detail: str,
        field: str = "",
        original_value: str = "",
        intermediate_value: str = "",
        final_value: str = "",
        transformation: str = "",
        result: str = "",
        severity: str = "INFO",
        evidence: str = "",
    ) -> AuditEvent:
        return AuditEvent(
            event_type=event_type,
            rule=rule,
            excel_row=_excel_row(row),
            document=normalize_text(_row_value(row, "DocumentoCliente", "Identification Number / CPF", "_documento_brasileiro")),
            account=normalize_text(_row_value(row, "NumConta", "AccountNumber")),
            person_type=normalize_text(_row_value(row, "Tipo de documento", "DocumentType")),
            action=action,
            detail=detail,
            field=field,
            original_value=original_value,
            intermediate_value=intermediate_value,
            final_value=final_value,
            transformation=transformation,
            result=result,
            severity=severity,
            evidence=evidence,
        )

    def _rule_event(
        self,
        row: dict[str, Any],
        rule: str,
        result: str,
        detail: str,
        severity: str,
        original_value: str = "",
        final_value: str = "",
        evidence: str = "",
        required_data: str = "",
        found_data: str = "",
    ) -> AuditEvent:
        return AuditEvent(
            event_type="AVALIACAO_REGRA",
            rule=rule,
            excel_row=_excel_row(row),
            document=normalize_text(_row_value(row, "DocumentoCliente", "Identification Number / CPF", "_documento_brasileiro")),
            account=normalize_text(_row_value(row, "NumConta", "AccountNumber")),
            person_type=normalize_text(_row_value(row, "Tipo de documento", "DocumentType")),
            action="avaliacao_regra",
            detail=detail,
            original_value=original_value,
            final_value=final_value,
            result=result,
            severity=severity,
            evidence=evidence,
            required_data=required_data,
            found_data=found_data,
        )

    def _summary_rule_event(
        self,
        rule: str,
        result: str,
        detail: str,
        severity: str,
        required_data: str = "",
        found_data: str = "",
    ) -> AuditEvent:
        return AuditEvent(
            event_type="AVALIACAO_REGRA",
            rule=rule,
            excel_row=None,
            action="avaliacao_regra",
            detail=detail,
            result=result,
            severity=severity,
            required_data=required_data,
            found_data=found_data,
        )

    def _issue_rows(self, issues: list[ValidationIssue]) -> list[dict[str, object]]:
        return [
            {
                "tipo": issue.level.upper(),
                "regra": issue.code,
                "linha_excel": issue.excel_row or "",
                "documento_mascarado": "",
                "conta_mascarada": "",
                "tipo_pessoa": "",
                "acao": "erro_validacao",
                "detalhe": issue.message,
            }
            for issue in issues
        ]

    def _fatca_pending_row(self, event: AuditEvent) -> dict[str, object]:
        now = datetime.now().replace(microsecond=0).isoformat(sep=" ")
        return {
            "linha_planilha": event.excel_row or "",
            "cliente_mascarado": "",
            "documento_brasileiro_mascarado": mask_value(event.document),
            "numero_conta_mascarado": mask_value(event.account),
            "classificacao_fatca": "FATCA",
            "us_tax_id_informado": "nao",
            "status": "PENDING_FISCAL_REVIEW",
            "motivo_ausencia": "US Tax ID ausente; CPF/CNPJ brasileiro nao utilizado como US Tax ID.",
            "tratamento_aplicado": event.detail,
            "elemento_omitido_ou_gerado": "TIN omitido",
            "resultado_validacao_xsd": "validado_apos_geracao",
            "bloqueante_para_envio": "sim",
            "data_hora": now,
        }

    def _replace_sheet(
        self,
        workbook: openpyxl.Workbook,
        name: str,
        rows: list[dict[str, object]],
        progress_callback: ProgressCallback | None = None,
    ) -> None:
        if name in workbook.sheetnames:
            del workbook[name]
        sheet = workbook.create_sheet(name)
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook["Sheet"]
        if not rows:
            rows = [{"status": "sem_registros"}]
        headers = list(rows[0].keys())
        sheet.append(headers)
        widths = [len(str(header)) for header in headers]
        total = len(rows)
        self._emit_progress(progress_callback, f"Gerando auditoria: {name}", 0, total, name)
        for index, row in enumerate(rows, 1):
            values = [row.get(header, "") for header in headers]
            sheet.append(values)
            for column_index, value in enumerate(values):
                widths[column_index] = min(max(widths[column_index], len(str(value or ""))), 78)
            self._emit_progress(progress_callback, f"Gerando auditoria: {name}", index, total, f"{name} linha {index}", force=index == total)
        for column_index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(column_index)].width = min(width + 2, 80)

    def _csv_rows(self, prepared: PreparedData, results: list[Any], reports: dict[str, Any]) -> list[dict[str, object]]:
        result_by_kind = {getattr(result, "kind", ""): result for result in results}
        rows: list[dict[str, object]] = []
        prepared_by_line = self._prepared_by_line(prepared.rows)
        events_by_line = self._events_by_line(prepared.events)
        rule_results = self._rule_results_by_line(prepared.events)
        account_doc_refs = self._account_doc_refs(reports)
        for index, row in enumerate(prepared.original_rows, 1):
            line = _excel_row(row)
            normalized = prepared_by_line.get(line)
            line_events = events_by_line.get(line, [])
            account = normalize_text(_row_value(row, "NumConta", "AccountNumber"))
            document = normalize_text(_row_value(row, "DocumentoCliente", "Identification Number / CPF"))
            final_balance = _row_value(normalized or row, "SaldoTotal", "Saldo da conta em 31/12/2025")
            decision = "incluido" if normalized is not None else "excluido"
            rows.append(
                {
                    "identificador": f"REG-{index:05d}",
                    "linha_origem": _excel_row(row) or "",
                    "conta": mask_value(account),
                    "nome": mask_value(_row_value(row, "NomeCliente", "Name")),
                    "PF_PJ": _row_value(row, "Tipo de documento", "DocumentType"),
                    "documento_original": mask_value(document),
                    "documento_normalizado": mask_value(_row_value(normalized or row, "_documento_brasileiro", "DocumentoCliente", "Identification Number / CPF")),
                    "saldo_original": _row_value(row, "SaldoTotal", "Saldo da conta em 31/12/2025"),
                    "saldo_final": final_balance,
                    "status_CI": _row_value(row, "DataHoraEncerramento CI"),
                    "status_CC": _row_value(row, "Status em 31/12 em CC"),
                    "REGRA_01 resultado": rule_results.get((line, "REGRA_01"), ""),
                    "REGRA_02 resultado": rule_results.get((line, "REGRA_02"), ""),
                    "REGRA_03 resultado": rule_results.get((line, "REGRA_03"), ""),
                    "decisao": decision,
                    "regra_aplicada": ";".join(event.rule for event in line_events),
                    "CRS_incluido": "sim" if normalized is not None and "crs" in reports else "nao",
                    "FATCA_incluido": "sim" if normalized is not None and "fatca" in reports else "nao",
                    "DocRefId": account_doc_refs.get(account, ""),
                    "US_TIN_status": _row_value(normalized or row, "_fatca_us_tin_status"),
                    "politica_US_TIN": _row_value(normalized or row, "_fatca_us_tin_policy"),
                    "resultado_XSD": self._xsd_summary(result_by_kind),
                    "severidade": "BLOQUEIO" if prepared.issues else ("ALERTA" if line_events else "INFO"),
                    "mensagem": "; ".join(event.detail for event in line_events),
                }
            )
        if not rows:
            rows.append({"identificador": "", "linha_origem": "", "mensagem": "sem registros"})
        return rows

    def _summary_rows(self, prepared: PreparedData, profile: MappingProfile | None, excel_path: Path | None, file_hash: str, results: list[Any], reports: dict[str, Any]) -> list[dict[str, object]]:
        rows = prepared.rows
        all_rows = prepared.original_rows
        xsd = {getattr(result, "kind", ""): "valido" if getattr(result, "valid", False) else "invalido" for result in results}
        policy = _profile_fixed(profile, "fatca.missing_us_tin_policy", "")
        missing_us_tin = sum(1 for row in rows if not row.get("_fatca_us_tin"))
        transformations = self._transformation_rows(prepared)
        rule_rows = self._rule_evaluation_rows(prepared)
        controlling_records = [record for row in rows for record in _controlling_audit_records(row)]
        included_controlling = [record for record in controlling_records if record.result == "INCLUIDO"]
        controlling_by_account: dict[str, int] = {}
        for record in included_controlling:
            controlling_by_account[record.account_number] = controlling_by_account.get(record.account_number, 0) + 1
        cp_conciliation = self._controlling_conciliation_rows(prepared, reports)
        final_status = "APTO PARA ENVIO"
        if policy in {"TECHNICAL_TEST_ONLY", "PENDING_FISCAL_CONFIRMATION"} or missing_us_tin:
            final_status = "NAO APTO PARA ENVIO"
        if prepared.issues or any(not getattr(result, "valid", False) for result in results):
            final_status = "NAO APTO PARA ENVIO"
        if any(row.get("resultado") == "divergente" for row in cp_conciliation):
            final_status = "NAO APTO PARA ENVIO"
        return [
            {
                "nome do arquivo de origem": excel_path.name if excel_path else "",
                "hash SHA-256 do arquivo de origem": file_hash,
                "data/hora do processamento": datetime.now().replace(microsecond=0).isoformat(sep=" "),
                "versao do aplicativo": "0.1.0",
                "ambiente": f"{platform.system()} {platform.release()}",
                "periodo reportado": _profile_fixed(profile, "message.reporting_period", "2025-12-31"),
                "modo": "TESTE" if policy in {"TECHNICAL_TEST_ONLY", "PENDING_FISCAL_CONFIRMATION"} else "PRODUCAO",
                "politica de US TIN": policy,
                "total de registros recebidos": len(all_rows),
                "total considerado": len(rows),
                "total incluido no CRS": len(getattr(reports.get("crs"), "accounts", [])),
                "total incluido no FATCA": len(getattr(reports.get("fatca"), "accounts", [])),
                "total excluido": len(all_rows) - len(rows),
                "total bloqueado por erro": len(prepared.issues),
                "total com alerta": sum(1 for event in prepared.events if event.event_type in {"AJUSTE", "PENDENCIA_US_TAX_ID"}),
                "total de PF": sum(1 for row in rows if normalize_text(_row_value(row, "Tipo de documento", "DocumentType")).upper() == "PF"),
                "total de PJ": sum(1 for row in rows if normalize_text(_row_value(row, "Tipo de documento", "DocumentType")).upper() == "PJ"),
                "quantidade de saldos negativos zerados": sum(1 for row in transformations if row.get("transformacao") == "NEGATIVE_TO_ZERO"),
                "quantidade de CPFs normalizados": sum(1 for row in transformations if row.get("campo") == "cpf"),
                "quantidade de CNPJs normalizados": sum(1 for row in transformations if row.get("campo") == "cnpj"),
                "quantidade de documentos invalidos": sum(1 for issue in prepared.issues if issue.code.startswith("DOC")),
                "quantidade de US TINs ausentes": missing_us_tin,
                "empresas com controlling person": len(controlling_by_account),
                "total de controlling persons recebidos": len(controlling_records),
                "total de controlling persons incluidos": len(included_controlling),
                "total de controlling persons excluidos": len(controlling_records) - len(included_controlling),
                "CPFs de controlling persons normalizados": sum(1 for record in controlling_records if record.normalized and record.tin_issued_by == "BR"),
                "CPFs de controlling persons invalidos": sum(1 for issue in prepared.issues if issue.code == "CP001" and "CPF invalido" in issue.message),
                "maximo de controladores por empresa": max(controlling_by_account.values(), default=0),
                "blocos vazios de controlling person ignorados": sum(int((row.get("_controlling_person_metrics") or {}).get("empty_blocks", 0)) for row in rows),
                "regras nao avaliadas": sum(1 for row in rule_rows if row.get("resultado") == "NAO_AVALIADA_DADO_AUSENTE"),
                "resultado XSD CRS": xsd.get("crs", "nao gerado"),
                "resultado XSD FATCA": xsd.get("fatca", "nao gerado"),
                "conclusao geral": final_status,
                "processing_id": prepared.processing_id,
                "usuario": os.environ.get("USERNAME") or os.environ.get("USER") or "",
            }
        ]

    def _input_rows(self, prepared: PreparedData) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        prepared_by_line = self._prepared_by_line(prepared.rows)
        for index, row in enumerate(prepared.original_rows, 1):
            normalized = prepared_by_line.get(_excel_row(row))
            base = {
                "numero sequencial": index,
                "linha da planilha": _excel_row(row) or "",
                "identificador interno": f"REG-{index:05d}",
                "nome": mask_value(_row_value(row, "NomeCliente", "Name")),
                "tipo PF/PJ": _row_value(row, "Tipo de documento", "DocumentType"),
                "CPF/CNPJ original": mask_value(_row_value(row, "DocumentoCliente", "Identification Number / CPF")),
                "CPF/CNPJ normalizado": mask_value(_row_value(normalized or row, "_documento_brasileiro", "DocumentoCliente", "Identification Number / CPF")),
                "conta": mask_value(_row_value(row, "NumConta", "AccountNumber")),
                "saldo original": _row_value(row, "SaldoTotal", "Saldo da conta em 31/12/2025"),
                "moeda original": _row_value(row, "Currency", "Moeda"),
                "data de nascimento original": _row_value(row, "Data de Nascimento"),
                "data de encerramento CI": _row_value(row, "DataHoraEncerramento CI"),
                "status CC": _row_value(row, "Status em 31/12 em CC"),
                "data do status": _row_value(row, "Encerramento CC"),
            }
            rows.append(base)
        return rows

    def _decision_rows(self, prepared: PreparedData) -> list[dict[str, object]]:
        rows = []
        for event in prepared.events:
            if event.event_type == "AVALIACAO_REGRA":
                continue
            rows.append(
                {
                    "identificador do registro": f"L{event.excel_row or ''}",
                    "conta": mask_value(event.account),
                    "documento": mask_value(event.document),
                    "incluido ou excluido": "excluido" if event.event_type == "REMOVIDO" else "incluido_com_observacao",
                    "destino CRS": "nao" if event.event_type == "REMOVIDO" else "sim",
                    "destino FATCA": "nao" if event.event_type == "REMOVIDO" else "sim",
                    "regra aplicada": event.rule,
                    "codigo da regra": event.rule,
                    "descricao detalhada": event.detail,
                    "valores utilizados na decisao": event.action,
                    "data/hora": datetime.now().replace(microsecond=0).isoformat(sep=" "),
                    "severidade": "BLOQUEIO" if event.event_type == "REMOVIDO" else ("ALERTA" if event.event_type.startswith("PENDENCIA") else "INFO"),
                }
            )
        rows.extend({"identificador do registro": issue.excel_row or "", "conta": "", "documento": "", "incluido ou excluido": "bloqueado", "destino CRS": "nao", "destino FATCA": "nao", "regra aplicada": issue.code, "codigo da regra": issue.code, "descricao detalhada": issue.message, "valores utilizados na decisao": issue.field, "data/hora": datetime.now().replace(microsecond=0).isoformat(sep=" "), "severidade": "BLOQUEIO"} for issue in prepared.issues)
        return rows

    def _rule_evaluation_rows(self, prepared: PreparedData) -> list[dict[str, object]]:
        rows = []
        for event in prepared.events:
            if event.event_type != "AVALIACAO_REGRA":
                continue
            rows.append(
                {
                    "processing_id": prepared.processing_id,
                    "linha da origem": event.excel_row or "",
                    "registro": f"L{event.excel_row or ''}",
                    "conta": mask_value(event.account),
                    "documento": mask_value(event.document),
                    "regra": event.rule,
                    "dados necessarios": event.required_data,
                    "dados encontrados": event.found_data,
                    "resultado": event.result,
                    "decisao": event.action,
                    "motivo": event.detail,
                    "severidade": event.severity,
                    "evidencia": event.evidence,
                }
            )
        return rows

    def _exclusion_rows(self, prepared: PreparedData) -> list[dict[str, object]]:
        return [
            {
                "registro": event.excel_row or "",
                "conta": mask_value(event.account),
                "cliente": "",
                "documento": mask_value(event.document),
                "regra": event.rule,
                "motivo": event.detail,
                "valores de origem": event.action,
                "data relevante": "",
                "status relevante": "",
                "evidencia da decisao": event.detail,
            }
            for event in prepared.events
            if event.event_type == "REMOVIDO"
        ]

    def _transformation_rows(self, prepared: PreparedData) -> list[dict[str, object]]:
        rows = []
        for event in prepared.events:
            if event.event_type in {"AJUSTE", "PENDENCIA_US_TAX_ID"}:
                rows.append(
                    {
                        "processing_id": prepared.processing_id,
                        "registro": event.excel_row or "",
                        "conta": mask_value(event.account),
                        "campo": event.field,
                        "tipo de dado": event.person_type,
                        "valor original": event.original_value,
                        "valor intermediario": event.intermediate_value,
                        "valor final": event.final_value,
                        "transformacao": event.transformation or event.action,
                        "motivo": event.detail,
                        "regra": event.rule,
                        "severidade": event.severity,
                        "alerta associado": "sim" if event.event_type == "PENDENCIA_US_TAX_ID" else "nao",
                    }
                )
        return rows

    def _report_rows(self, report: Any, kind: str) -> list[dict[str, object]]:
        if not report:
            return []
        rows = []
        message_id = report.message_spec.message_ref_id
        for account in report.accounts:
            holder = account.account_holder
            tax = holder.tins[0] if holder and holder.tins else None
            rows.append(
                {
                    "numero da conta" if kind == "crs" else "conta": mask_value(account.account_number),
                    "CPF/CNPJ" if kind == "crs" else "cliente": mask_value(getattr(holder, "documento_brasileiro", "")),
                    "tipo de titular" if kind == "crs" else "tipo": holder.kind if holder else "",
                    "MessageRefId": message_id,
                    "DocRefId": account.doc_spec.doc_ref_id,
                    "pais de residencia" if kind == "crs" else "US TIN": ",".join(holder.res_country_codes) if kind == "crs" and holder else getattr(holder, "fatca_us_tin", ""),
                    "elemento fiscal utilizado" if kind == "crs" else "status do US TIN": "TIN" if kind == "crs" and holder else getattr(holder, "fatca_us_tin_status", ""),
                    "valor fiscal enviado" if kind == "crs" else "motivo da ausencia": tax.value if tax else "",
                    "issuedBy" if kind == "crs" else "politica aplicada": tax.issued_by if kind == "crs" and tax else getattr(holder, "fatca_us_tin_policy", ""),
                    "saldo": account.account_balance,
                    "moeda": account.account_currency,
                    "conta encerrada": account.closed_account,
                    "nascimento" if kind == "crs" else "apto para envio": getattr(holder, "birth_date", "") if kind == "crs" else ("nao" if getattr(holder, "fatca_us_tin_blocking", "") == "sim" or getattr(holder, "fatca_us_tin_policy", "") == "TECHNICAL_TEST_ONLY" else "sim"),
                    "resultado de validacoes internas" if kind == "crs" else "pendencia fiscal": "ok" if kind == "crs" else getattr(holder, "fatca_us_tin_reason", ""),
                }
            )
        return rows

    def _controlling_person_rows(self, prepared: PreparedData, reports: dict[str, Any]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for row in prepared.rows:
            account = normalize_text(_row_value(row, "NumConta", "AccountNumber"))
            doc_ref_id = self._account_doc_ref({"crs": reports.get("crs")}, account)
            for record in _controlling_audit_records(row):
                rows.append(record.audit_row(prepared.processing_id, doc_ref_id))
        return rows

    def _controlling_conciliation_rows(self, prepared: PreparedData, reports: dict[str, Any]) -> list[dict[str, object]]:
        crs_accounts = {account.account_number: account for account in getattr(reports.get("crs"), "accounts", [])}
        rows: list[dict[str, object]] = []
        expected_total = 0
        generated_total = 0
        for row in prepared.rows:
            account_number = normalize_text(_row_value(row, "NumConta", "AccountNumber"))
            expected = len(_included_controlling_records(row))
            generated = len(getattr(crs_accounts.get(account_number), "controlling_persons", []))
            expected_total += expected
            generated_total += generated
            if expected or generated:
                rows.append(
                    {
                        "linha da origem": _excel_row(row) or "",
                        "conta": mask_value(account_number),
                        "empresa": _row_value(row, "Name", "NomeCliente"),
                        "documento da empresa": mask_value(_row_value(row, "_documento_brasileiro", "Identification Number / CPF", "DocumentoCliente")),
                        "controladores validos na planilha": expected,
                        "ControllingPerson no CRS": generated,
                        "resultado": "ok" if expected == generated else "divergente",
                        "motivo": "" if expected == generated else "Quantidade de controladores no XML diferente da planilha.",
                    }
                )
        rows.append(
            {
                "linha da origem": "",
                "conta": "",
                "empresa": "TOTAL",
                "documento da empresa": "",
                "controladores validos na planilha": expected_total,
                "ControllingPerson no CRS": generated_total,
                "resultado": "ok" if expected_total == generated_total else "divergente",
                "motivo": "" if expected_total == generated_total else "Total de ControllingPerson no CRS diverge da planilha.",
            }
        )
        return rows

    def _identifier_rows(self, reports: dict[str, Any]) -> list[dict[str, object]]:
        rows = []
        seen: set[str] = set()
        for kind, report in reports.items():
            ids = [("MessageRefId", "mensagem", report.message_spec.message_ref_id), ("DocRefId", "ReportingFI", report.reporting_fi.doc_spec.doc_ref_id)]
            ids.extend(("DocRefId", account.account_number, account.doc_spec.doc_ref_id) for account in report.accounts)
            for id_type, entity, value in ids:
                duplicate = value in seen
                seen.add(value)
                rows.append({"tipo do identificador": id_type, "entidade relacionada": f"{kind}:{entity}", "identificador": value, "regra de formacao": "perfil DITC configurado", "componentes utilizados": value, "tamanho": len(value), "resultado da validacao": "duplicado" if duplicate else "ok", "indicador de duplicidade": "sim" if duplicate else "nao"})
        return rows

    def _xsd_rows(self, results: list[Any]) -> list[dict[str, object]]:
        rows = []
        for result in results:
            issues = getattr(result, "issues", [])
            if not issues:
                rows.append({"tipo de arquivo": result.kind, "nome do XML": result.xml_path, "XSD utilizado": result.kind, "valido": "sim", "quantidade de erros": 0, "severidade": "INFO", "linha": "", "coluna": "", "XPath": "", "mensagem": "XML valido contra XSD.", "schema ou namespace relacionado": ""})
            for issue in issues:
                rows.append({"tipo de arquivo": result.kind, "nome do XML": result.xml_path, "XSD utilizado": result.kind, "valido": "nao", "quantidade de erros": len(issues), "severidade": issue.level.upper(), "linha": issue.excel_row or "", "coluna": "", "XPath": issue.field, "mensagem": issue.message, "schema ou namespace relacionado": issue.code})
        return rows

    def _conciliation_rows(self, prepared: PreparedData, reports: dict[str, Any]) -> list[dict[str, object]]:
        crs_accounts = {account.account_number: account for account in getattr(reports.get("crs"), "accounts", [])}
        fatca_accounts = {account.account_number: account for account in getattr(reports.get("fatca"), "accounts", [])}
        rows = []
        prepared_lines = set(self._prepared_by_line(prepared.rows))
        for row in prepared.original_rows:
            account = normalize_text(_row_value(row, "NumConta", "AccountNumber"))
            expected = _excel_row(row) in prepared_lines
            rows.append({"estava na entrada": "sim", "deveria ser incluido": "sim" if expected else "nao", "apareceu no CRS": "sim" if account in crs_accounts else "nao", "apareceu no FATCA": "sim" if account in fatca_accounts else "nao", "quantidade de ocorrencias": int(account in crs_accounts) + int(account in fatca_accounts), "conta correspondente": mask_value(account), "documento correspondente": mask_value(_row_value(row, "DocumentoCliente", "Identification Number / CPF")), "divergencia encontrada": "nao" if (expected == (account in crs_accounts or account in fatca_accounts)) else "sim"})
        return rows

    def _pending_rows(self, prepared: PreparedData) -> list[dict[str, object]]:
        rows = [self._fatca_pending_row(event) for event in prepared.events if event.event_type == "PENDENCIA_US_TAX_ID"]
        rows.extend({"linha_planilha": issue.excel_row or "", "cliente_mascarado": "", "documento_brasileiro_mascarado": "", "numero_conta_mascarado": "", "classificacao_fatca": "", "us_tax_id_informado": "", "status": "BLOQUEIO", "motivo_ausencia": issue.message, "tratamento_aplicado": issue.suggestion, "elemento_omitido_ou_gerado": "", "resultado_validacao_xsd": "", "bloqueante_para_envio": "sim", "data_hora": datetime.now().replace(microsecond=0).isoformat(sep=" ")} for issue in prepared.issues)
        rows.append({"linha_planilha": "", "cliente_mascarado": "", "documento_brasileiro_mascarado": "", "numero_conta_mascarado": "", "classificacao_fatca": "CRS", "us_tax_id_informado": "", "status": "CRS_PJ_TIN", "motivo_ausencia": "Titular PJ CRS usa CNPJ em TIN issuedBy=BR.", "tratamento_aplicado": "ReportingFI CRS permanece com IN issuedBy=KY; AccountHolder/Organisation PJ usa TIN issuedBy=BR.", "elemento_omitido_ou_gerado": "TIN", "resultado_validacao_xsd": "validado_apos_geracao", "bloqueante_para_envio": "nao", "data_hora": datetime.now().replace(microsecond=0).isoformat(sep=" ")})
        return rows

    def _hash_rows(self, prepared: PreparedData, excel_path: Path | None, file_hash: str, results: list[Any], csv_path: Path, xlsx_path: Path, json_path: Path) -> list[dict[str, object]]:
        paths = []
        if excel_path:
            paths.append(excel_path)
        paths.extend(Path(result.xml_path) for result in results if getattr(result, "xml_path", ""))
        paths.extend(path for path in (csv_path, xlsx_path, json_path) if path.name)
        rows = []
        for path in paths:
            rows.append(
                {
                    "processing_id": prepared.processing_id,
                    "nome": path.name,
                    "tamanho": path.stat().st_size if path.exists() else "",
                    "SHA-256": file_hash if excel_path and path.resolve() == excel_path.resolve() else (sha256_file(path) if path.exists() else ""),
                    "data_criacao": datetime.fromtimestamp(path.stat().st_mtime).isoformat(sep=" ") if path.exists() else "",
                }
            )
        return rows

    def _manifest(
        self,
        prepared: PreparedData,
        profile: MappingProfile | None,
        excel_path: Path | None,
        file_hash: str,
        results: list[Any],
        reports: dict[str, Any],
        compact: bool = False,
    ) -> dict[str, Any]:
        manifest_prepared = self._sample_prepared(prepared, AUDIT_XLSX_SAMPLE_LIMIT) if compact else prepared
        return {
            "processing_id": prepared.processing_id,
            "application_version": "0.1.0",
            "source_file": {"name": excel_path.name if excel_path else "", "sha256": file_hash},
            "reporting_period": _profile_fixed(profile, "message.reporting_period", "2025-12-31"),
            "mode": "TESTE" if _profile_fixed(profile, "fatca.missing_us_tin_policy", "") in {"TECHNICAL_TEST_ONLY", "PENDING_FISCAL_CONFIRMATION"} else "PRODUCAO",
            "generated_at": datetime.now().replace(microsecond=0).isoformat(),
            "environment": f"{platform.system()} {platform.release()}",
            "user": os.environ.get("USERNAME") or os.environ.get("USER") or "",
            "counts": self._summary_rows(prepared, profile, excel_path, file_hash, results, reports)[0],
            "identifier_profile": "DITC_SEQUENCE" if profile and not profile.identifier_config.use_uuid else "INTERNAL_TEST_UUID",
            "us_tin_policy": _profile_fixed(profile, "fatca.missing_us_tin_policy", ""),
            "rules": {"REGRA_01": "encerramento CI antes do periodo", "REGRA_02": "menor conta duplicada", "REGRA_03": "status CC encerrada no periodo", "FATCA_US_TIN": "US TIN separado do CPF/CNPJ"},
            "manifest_detail_mode": "amostra" if compact else "completo",
            "manifest_sample_limit": AUDIT_XLSX_SAMPLE_LIMIT if compact else "",
            "rule_evaluations": self._rule_evaluation_rows(manifest_prepared),
            "excluded_records": self._exclusion_rows(manifest_prepared),
            "transformations": self._transformation_rows(manifest_prepared),
            "identifiers": self._identifier_summary_rows(reports) if compact else self._identifier_rows(reports),
            "xsd_validation": self._xsd_rows(results),
            "reconciliation": self._conciliation_rows(manifest_prepared, reports),
            "controlling_persons": self._controlling_person_rows(manifest_prepared, reports),
            "controlling_person_reconciliation": self._controlling_conciliation_rows(manifest_prepared, reports),
            "pending_decisions": self._pending_rows(manifest_prepared),
            "file_hashes": self._hash_rows(prepared, excel_path, file_hash, results, Path(""), Path(""), Path("")),
            "final_status": self._summary_rows(prepared, profile, excel_path, file_hash, results, reports)[0]["conclusao geral"],
        }

    def _prepared_by_line(self, rows: list[dict[str, Any]]) -> dict[int | None, dict[str, Any]]:
        return {_excel_row(row): row for row in rows}

    def _events_by_line(self, events: list[AuditEvent]) -> dict[int | None, list[AuditEvent]]:
        indexed: dict[int | None, list[AuditEvent]] = {}
        for event in events:
            indexed.setdefault(event.excel_row, []).append(event)
        return indexed

    def _rule_results_by_line(self, events: list[AuditEvent]) -> dict[tuple[int | None, str], str]:
        indexed: dict[tuple[int | None, str], str] = {}
        for event in events:
            if event.event_type == "AVALIACAO_REGRA":
                indexed[(event.excel_row, event.rule)] = event.result
        return indexed

    def _account_doc_refs(self, reports: dict[str, Any]) -> dict[str, str]:
        refs: dict[str, str] = {}
        for report in reports.values():
            for account in getattr(report, "accounts", []):
                refs[normalize_text(account.account_number)] = account.doc_spec.doc_ref_id
        return refs

    def _identifier_summary_rows(self, reports: dict[str, Any]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for kind, report in reports.items():
            rows.append({"tipo de arquivo": kind, "tipo do identificador": "MessageRefId", "quantidade": 1, "exemplo": report.message_spec.message_ref_id})
            rows.append({"tipo de arquivo": kind, "tipo do identificador": "ReportingFI DocRefId", "quantidade": 1, "exemplo": report.reporting_fi.doc_spec.doc_ref_id})
            first_account = next(iter(getattr(report, "accounts", [])), None)
            rows.append(
                {
                    "tipo de arquivo": kind,
                    "tipo do identificador": "Account DocRefId",
                    "quantidade": len(getattr(report, "accounts", [])),
                    "exemplo": first_account.doc_spec.doc_ref_id if first_account else "",
                }
            )
        return rows

    def _matching_prepared_row(self, original: dict[str, Any], rows: list[dict[str, Any]]) -> dict[str, Any] | None:
        original_line = _excel_row(original)
        for row in rows:
            if _excel_row(row) == original_line:
                return row
        return None

    def _rule_for(self, row: dict[str, Any], events: list[AuditEvent]) -> str:
        line = _excel_row(row)
        return ";".join(event.rule for event in events if event.excel_row == line)

    def _rule_result_for(self, row: dict[str, Any], events: list[AuditEvent], rule: str) -> str:
        line = _excel_row(row)
        match = next((event for event in events if event.excel_row == line and event.rule == rule and event.event_type == "AVALIACAO_REGRA"), None)
        return match.result if match else ""

    def _message_for(self, row: dict[str, Any], events: list[AuditEvent]) -> str:
        line = _excel_row(row)
        return "; ".join(event.detail for event in events if event.excel_row == line)

    def _account_doc_ref(self, reports: dict[str, Any], account_number: str) -> str:
        for report in reports.values():
            for account in getattr(report, "accounts", []):
                if normalize_text(account.account_number) == normalize_text(account_number):
                    return account.doc_spec.doc_ref_id
        return ""

    def _xsd_summary(self, result_by_kind: dict[str, Any]) -> str:
        return ";".join(f"{kind}:{'valido' if result.valid else 'invalido'}" for kind, result in result_by_kind.items())


def _column_for(profile: MappingProfile, field: str, fallback: str) -> str:
    rule = profile.field_mappings.get(field)
    if rule and rule.source == "column" and rule.column:
        return rule.column
    return fallback


def _row_value(row: dict[str, Any] | None, *names: str) -> Any:
    if not row:
        return ""
    for name in names:
        if name in row and row.get(name) not in (None, ""):
            return row.get(name)
    return ""


def _profile_fixed(profile: MappingProfile | None, field: str, default: str) -> str:
    if not profile:
        return default
    rule = profile.field_mappings.get(field)
    if rule and rule.source == "fixed" and rule.fixed_value:
        return rule.fixed_value
    return default


def _controlling_audit_records(row: dict[str, Any]) -> list[ControllingPersonRecord]:
    records = row.get("_controlling_person_audit") or []
    return [record for record in records if isinstance(record, ControllingPersonRecord)]


def _included_controlling_records(row: dict[str, Any]) -> list[ControllingPersonRecord]:
    records = row.get("_controlling_persons") or []
    return [record for record in records if isinstance(record, ControllingPersonRecord)]


def _reporting_start(profile: MappingProfile) -> date:
    reporting_period = _profile_fixed(profile, "message.reporting_period", "2025-12-31")
    try:
        parsed = datetime.fromisoformat(to_date(reporting_period)).date()
    except ValueError:
        parsed = date(2025, 12, 31)
    return date(parsed.year, 1, 1)


def _normalize_status(value: Any) -> str:
    import unicodedata

    text = unicodedata.normalize("NFKD", normalize_text(value).casefold())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.split())


def _has_column(rows: list[dict[str, Any]], column: str) -> bool:
    return any(column in row for row in rows)


def _excel_row(row: dict[str, Any]) -> int | None:
    value = row.get("_excel_row")
    return int(value) if isinstance(value, int) else None


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if is_empty(value):
        return None
    text = normalize_text(value)
    if text.casefold() == "null":
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _decimal(value: Any) -> Decimal | None:
    if is_empty(value):
        return None
    text = normalize_text(value)
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _account_sort_key(value: Any) -> tuple[int, int | str]:
    text = normalize_text(value)
    digits = "".join(char for char in text if char.isdigit())
    if digits and digits == text:
        return (0, int(digits))
    return (1, text)
