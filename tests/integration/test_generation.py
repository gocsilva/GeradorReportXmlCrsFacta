from __future__ import annotations

import json
import os
import re
import shutil
from pathlib import Path

import pytest
import openpyxl
from lxml import etree

from crs_fatca_generator.app import build_sample_profile
from crs_fatca_generator.app import generate_from_excel
from crs_fatca_generator.infrastructure import paths as path_config
from crs_fatca_generator.infrastructure.paths import default_crs_schema, default_fatca_schema
from crs_fatca_generator.models.mapping import MappingRule
from crs_fatca_generator.services.excel_reader import ExcelReader
from crs_fatca_generator.services.generation_service import GenerationService
from crs_fatca_generator.services.mapping_service import MappingService, infer_default_profile, missing_simple_columns, simple_output_paths
from crs_fatca_generator.services.schema_inspector import SchemaInspector
from crs_fatca_generator.services.data_preparation_service import DataPreparationService
from crs_fatca_generator.services.xml_validator import XmlValidator
from crs_fatca_generator.services.xml_helpers import CRS_NS
from crs_fatca_generator.services.controlling_person_service import detect_controlling_person_blocks, extract_controlling_persons
from crs_fatca_generator.services.xml_splitter_service import XmlSplitterService
import crs_fatca_generator.services.data_preparation_service as preparation_module
import crs_fatca_generator.services.xml_splitter_service as splitter_module


def exemplos_excel_path() -> Path:
    for path in (Path("ExemplosDados/schema_mock.xlsx"), Path("ExemplosDados/mock_dados_teste.xlsx"), Path("ExemplosDados/novo_layout_dados_mock.xlsx")):
        if path.exists():
            return path
    return Path("ExemplosDados/novo_layout_dados_mock.xlsx")


def canonical(path: Path) -> bytes:
    parser = etree.XMLParser(remove_blank_text=True)
    root = etree.parse(str(path), parser).getroot()
    for elem in root.iter():
        if elem.text and elem.text.startswith("AUTO-"):
            elem.text = "AUTO-ID"
    return etree.tostring(root, method="c14n")


def controlling_headers(blocks: int = 4) -> list[str]:
    headers = [
        "DocumentType",
        "Identification Number / CPF",
        "Name",
        "AccountNumber",
        "USPerson",
        "FirstName",
        "MiddleName",
        "LastName",
        "País de Residencia fiscal",
        "Account Number Type",
        "Closed Account?",
        "Data de Nascimento",
        "Account Holder Type",
        "Name Type",
        "Tax Residence",
        "TIN Issued by",
        "Address Type",
        "Country",
        "Street",
        "City",
        "Saldo da conta em 31/12/2025",
        "Currency",
        "Document Type",
    ]
    block = ["Name Type", "First Name", "Last Name", "Controlling Person Type", "Tax Residence", "Identification Number", "TIN Issued By", "Address Type", "Country", "City", "Birth Date"]
    for _ in range(blocks):
        headers.extend(block)
    return headers


def row_from_raw(headers: list[str], values: list[object]) -> dict[str, object]:
    row = {header if header not in headers[:idx] else f"{header} #{headers[:idx].count(header) + 1}": values[idx] if idx < len(values) else "" for idx, header in enumerate(headers)}
    row["_headers"] = headers
    row["_raw_values"] = values
    row["_excel_row"] = 2
    return row


def test_gera_crs_e_fatca_validos(tmp_path: Path) -> None:
    profile = build_sample_profile(tmp_path)
    service = GenerationService(default_crs_schema(), default_fatca_schema())
    results = service.generate(["crs", "fatca"], [{"_excel_row": 2}], profile, overwrite=True)
    assert [result.valid for result in results] == [True, True]
    assert (tmp_path / "CRS_teste.xml").exists()
    assert (tmp_path / "FATCA_teste.xml").exists()


def test_fatca_usa_apenas_linhas_usperson_true_quando_coluna_existe(tmp_path: Path) -> None:
    headers = ["DocumentoCliente", "Tipo de documento", "NumConta", "NomeCliente", "SaldoTotal", "Endereco", "Cidade", "Pais", "USPerson"]
    profile = infer_default_profile(headers)
    profile.output.fatca_path = str(tmp_path / "fatca_usperson.xml")
    rows = [
        {"DocumentoCliente": "06360698501", "Tipo de documento": "PF", "NumConta": "ACC-TRUE", "NomeCliente": "Cliente True", "SaldoTotal": "10", "Endereco": "Rua A", "Cidade": "Sao Paulo", "Pais": "BR", "USPerson": "true", "_excel_row": 2},
        {"DocumentoCliente": "09030562595", "Tipo de documento": "PF", "NumConta": "ACC-FALSE", "NomeCliente": "Cliente False", "SaldoTotal": "20", "Endereco": "Rua B", "Cidade": "Rio", "Pais": "BR", "USPerson": "false", "_excel_row": 3},
        {"DocumentoCliente": "16011329721", "Tipo de documento": "PF", "NumConta": "ACC-SIM", "NomeCliente": "Cliente Sim", "SaldoTotal": "30", "Endereco": "Rua C", "Cidade": "Curitiba", "Pais": "BR", "USPerson": "sim", "_excel_row": 4},
    ]

    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["fatca"], rows, profile, Path("entrada.xlsx"), overwrite=True)[0]

    assert result.valid is True
    assert result.summary["filtro_fatca_usperson"] == "aplicado"
    assert result.summary["linhas_fatca_usadas"] == 2
    assert result.summary["linhas_fatca_ignoradas_por_usperson"] == 1
    tree = etree.parse(str(tmp_path / "fatca_usperson.xml"))
    accounts = [item.text for item in tree.xpath(".//*[local-name()='AccountNumber']")]
    assert accounts == ["ACC-TRUE", "ACC-SIM"]


def test_geracao_crs_respeita_limite_de_tamanho_sem_quebrar_account_report(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(splitter_module, "_mb_to_bytes", lambda _value: 1800)
    profile = infer_default_profile(["DocumentoCliente", "Tipo de documento", "NumConta", "NomeCliente", "SaldoTotal", "Endereco", "Cidade", "Pais"])
    profile.output.crs_path = str(tmp_path / "crs_limitado.xml")
    profile.output.crs_size_limit_mb = 1
    rows = [
        {"DocumentoCliente": "06360698501", "Tipo de documento": "PF", "NumConta": "ACC-1", "NomeCliente": "Cliente Um", "SaldoTotal": "10", "Endereco": "Rua A", "Cidade": "Sao Paulo", "Pais": "BR", "_excel_row": 2},
        {"DocumentoCliente": "09030562595", "Tipo de documento": "PF", "NumConta": "ACC-2", "NomeCliente": "Cliente Dois", "SaldoTotal": "20", "Endereco": "Rua B", "Cidade": "Rio", "Pais": "BR", "_excel_row": 3},
        {"DocumentoCliente": "16011329721", "Tipo de documento": "PF", "NumConta": "ACC-3", "NomeCliente": "Cliente Tres", "SaldoTotal": "30", "Endereco": "Rua C", "Cidade": "Curitiba", "Pais": "BR", "_excel_row": 4},
    ]

    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["crs"], rows, profile, Path("entrada.xlsx"), overwrite=True)[0]

    paths = [Path(item.strip()) for item in result.xml_path.split(";")]
    assert result.valid is True
    assert len(paths) > 1
    assert not (tmp_path / "crs_limitado.xml").exists()
    assert sum(int(etree.parse(str(path)).xpath("count(.//*[local-name()='AccountReport'])")) for path in paths) == 3
    assert all(XmlValidator().validate_file(path, default_crs_schema(), "crs") == [] for path in paths)


def test_divide_xml_existente_por_account_report(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    profile = infer_default_profile(["DocumentoCliente", "Tipo de documento", "NumConta", "NomeCliente", "SaldoTotal", "Endereco", "Cidade", "Pais"])
    profile.output.crs_path = str(tmp_path / "crs_original.xml")
    rows = [
        {"DocumentoCliente": "06360698501", "Tipo de documento": "PF", "NumConta": "ACC-1", "NomeCliente": "Cliente Um", "SaldoTotal": "10", "Endereco": "Rua A", "Cidade": "Sao Paulo", "Pais": "BR", "_excel_row": 2},
        {"DocumentoCliente": "09030562595", "Tipo de documento": "PF", "NumConta": "ACC-2", "NomeCliente": "Cliente Dois", "SaldoTotal": "20", "Endereco": "Rua B", "Cidade": "Rio", "Pais": "BR", "_excel_row": 3},
        {"DocumentoCliente": "16011329721", "Tipo de documento": "PF", "NumConta": "ACC-3", "NomeCliente": "Cliente Tres", "SaldoTotal": "30", "Endereco": "Rua C", "Cidade": "Curitiba", "Pais": "BR", "_excel_row": 4},
    ]
    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["crs"], rows, profile, Path("entrada.xlsx"), overwrite=True)[0]
    assert result.valid is True

    monkeypatch.setattr(splitter_module, "_mb_to_bytes", lambda _value: 1800)
    parts = XmlSplitterService().split_existing_xml(tmp_path / "crs_original.xml", tmp_path / "partes", 1)

    assert len(parts) > 1
    assert sum(int(etree.parse(str(path)).xpath("count(.//*[local-name()='AccountReport'])")) for path in parts) == 3
    assert all(XmlValidator().validate_file(path, default_crs_schema(), "crs") == [] for path in parts)


def test_golden_files_validam_contra_xsd_real() -> None:
    validator = XmlValidator()
    assert validator.validate_file(Path("tests/golden_files/CRS_esperado.xml"), default_crs_schema(), "crs") == []
    assert validator.validate_file(Path("tests/golden_files/FATCA_esperado.xml"), default_fatca_schema(), "fatca") == []


def test_nil_report_fatca_valido(tmp_path: Path) -> None:
    profile = build_sample_profile(tmp_path)
    profile.field_mappings["nil_report.enabled"] = MappingRule("fixed", fixed_value="sim")
    results = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["fatca"], [{"_excel_row": 2}], profile, overwrite=True)
    assert results[0].valid is True
    assert "NilReport" in (tmp_path / "FATCA_teste.xml").read_text(encoding="utf-8")


def test_fatca_nil_report_por_checkbox_ignora_account_report(tmp_path: Path) -> None:
    profile = build_sample_profile(tmp_path)
    profile.output.fatca_nil_report = True
    row = {
        "DocumentoCliente": "06360698501",
        "Tipo de documento": "PF",
        "NumConta": "ACC-1",
        "NomeCliente": "Cliente Um",
        "SaldoTotal": "10",
        "Endereco": "Rua A",
        "Cidade": "Sao Paulo",
        "Pais": "BR",
        "_excel_row": 2,
    }

    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["fatca"], [row], profile, overwrite=True)[0]

    assert result.valid is True
    tree = etree.parse(str(tmp_path / "FATCA_teste.xml"))
    assert tree.xpath("count(.//*[local-name()='NilReport'])") == 1
    assert tree.xpath("count(.//*[local-name()='AccountReport'])") == 0
    assert tree.xpath("count(.//*[local-name()='ReportingFI'])") == 1
    assert tree.xpath("count(.//*[local-name()='Warning'])") == 0


def test_crs_controlling_person_valida(tmp_path: Path) -> None:
    profile = build_sample_profile(tmp_path)
    profile.field_mappings["holder.kind"] = MappingRule("fixed", fixed_value="organisation")
    profile.field_mappings["holder.organisation_name"] = MappingRule("fixed", fixed_value="Empresa XPTO")
    profile.field_mappings["holder.acct_holder_type"] = MappingRule("fixed", fixed_value="CRS101")
    profile.field_mappings["controlling.first_name"] = MappingRule("fixed", fixed_value="Maria")
    profile.field_mappings["controlling.last_name"] = MappingRule("fixed", fixed_value="Souza")
    profile.field_mappings["controlling.res_country"] = MappingRule("fixed", fixed_value="BR")
    profile.field_mappings["controlling.address_country"] = MappingRule("fixed", fixed_value="BR")
    profile.field_mappings["controlling.address_free"] = MappingRule("fixed", fixed_value="Rua C 3")
    profile.field_mappings["controlling.type"] = MappingRule("fixed", fixed_value="CRS801")
    profile.field_mappings["controlling.self_cert"] = MappingRule("fixed", fixed_value="CRS1001")
    results = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["crs"], [{"_excel_row": 2}], profile, overwrite=True)
    assert results[0].valid is True
    assert "ControllingPerson" in (tmp_path / "CRS_teste.xml").read_text(encoding="utf-8")


def test_fatca_substantial_owner_valida(tmp_path: Path) -> None:
    profile = build_sample_profile(tmp_path)
    profile.field_mappings["substantial.first_name"] = MappingRule("fixed", fixed_value="Carlos")
    profile.field_mappings["substantial.last_name"] = MappingRule("fixed", fixed_value="Lima")
    profile.field_mappings["substantial.res_country"] = MappingRule("fixed", fixed_value="BR")
    profile.field_mappings["substantial.address_country"] = MappingRule("fixed", fixed_value="BR")
    profile.field_mappings["substantial.address_free"] = MappingRule("fixed", fixed_value="Rua D 4")
    results = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["fatca"], [{"_excel_row": 2}], profile, overwrite=True)
    assert results[0].valid is True
    assert "SubstantialOwner" in (tmp_path / "FATCA_teste.xml").read_text(encoding="utf-8")


def test_exemplos_dados_autoconfiguram_e_validam(tmp_path: Path) -> None:
    excel_path = exemplos_excel_path()
    reader = ExcelReader()
    preview = reader.preview(excel_path, "Planilha1", 1)
    profile = infer_default_profile(preview.headers)
    profile.output.crs_path = str(tmp_path / "exemplos_crs.xml")
    profile.output.fatca_path = str(tmp_path / "exemplos_fatca.xml")
    rows = reader.read_rows(excel_path, "Planilha1", 1)
    results = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["crs", "fatca"], rows, profile, excel_path, overwrite=True)
    assert [result.valid for result in results] == [True, True]
    if excel_path.name in {"mock_dados_teste.xlsx", "novo_layout_dados_mock.xlsx"}:
        assert profile.grouping.account_key == "AccountNumber"
        assert profile.field_mappings["holder.kind"].column == "DocumentType"
        assert profile.field_mappings["holder.tin"].column == "Identification Number / CPF"
        assert profile.field_mappings["account.doc_type_indic"].transformations == ["code_prefix"]
        assert profile.field_mappings["account.currency"].column == "Currency"
    else:
        assert profile.grouping.account_key == "NumConta"
        assert profile.field_mappings["account.currency"].fixed_value == "USD"
    assert profile.field_mappings["holder.kind"].transformations == ["pf_pj_kind"]
    assert profile.field_mappings["holder.address_free"].source == "calculated"
    crs = etree.parse(str(tmp_path / "exemplos_crs.xml"))
    fatca = etree.parse(str(tmp_path / "exemplos_fatca.xml"))
    assert crs.findtext(".//{urn:oecd:ties:crs:v3}TransmittingCountry") == "KY"
    assert crs.findtext(".//{urn:oecd:ties:crs:v3}ReceivingCountry") == "BR"
    assert fatca.findtext(".//{urn:oecd:ties:stffatcatypes:v2}TransmittingCountry") == "KY"
    assert fatca.findtext(".//{urn:oecd:ties:stffatcatypes:v2}ReceivingCountry") == "US"
    assert crs.findtext(".//{urn:oecd:ties:crs:v3}MessageRefId").startswith("KY2025BRFI107442")
    assert "-" not in crs.findtext(".//{urn:oecd:ties:crs:v3}MessageRefId")
    crs_doc_refs = [item.text for item in crs.findall(".//{urn:oecd:ties:crsstf:v5}DocRefId")]
    assert len(crs_doc_refs) == len(set(crs_doc_refs))
    assert all(item.get("currCode") == "USD" for item in crs.findall(".//{urn:oecd:ties:crs:v3}AccountBalance"))
    assert all(item.get("currCode") == "USD" for item in fatca.findall(".//{urn:oecd:ties:fatca:v2}AccountBalance"))
    assert all(not (item.text or "").startswith("-") for item in crs.findall(".//{urn:oecd:ties:crs:v3}AccountBalance"))
    assert (tmp_path / f"{excel_path.stem}_relatorio_auditoria.csv").exists()
    assert (tmp_path / f"{excel_path.stem}_relatorio_auditoria.xlsx").exists()


def test_fluxo_botao_simples_preserva_perfil_ditc_e_gera_arquivos(tmp_path: Path) -> None:
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    qtwidgets = pytest.importorskip("PySide6.QtWidgets")
    qtcore = pytest.importorskip("PySide6.QtCore")
    from crs_fatca_generator.gui.main_window import MainWindow

    excel_path = Path("ExemplosDados/novo_layout_dados_mock.xlsx")
    assert excel_path.exists()
    app = qtwidgets.QApplication.instance() or qtwidgets.QApplication([])
    window = MainWindow()
    try:
        window.excel_path = excel_path
        window.excel_path_edit.setText(str(excel_path))
        window.sheet_combo.clear()
        window.sheet_combo.addItems(window.excel_reader.list_sheets(excel_path))
        window.load_preview()
        window.crs_output_edit.setText(str(tmp_path / "botao_CRS.xml"))
        window.fatca_output_edit.setText(str(tmp_path / "botao_FATCA.xml"))

        def configure_test_outputs() -> None:
            window.crs_output_edit.setText(str(tmp_path / "botao_CRS.xml"))
            window.fatca_output_edit.setText(str(tmp_path / "botao_FATCA.xml"))

        window.configure_simple_outputs = configure_test_outputs  # type: ignore[method-assign]
        button = window.findChild(qtwidgets.QPushButton, "execute_now_button")
        assert button is not None
        button.click()
        assert window.worker is not None
        assert window.worker.profile.identifier_config.prefix == "KY2025BRFI107442"
        assert window.worker.profile.identifier_config.use_uuid is False

        loop = qtcore.QEventLoop()
        timer = qtcore.QTimer()
        timer.setSingleShot(True)
        timer.timeout.connect(loop.quit)
        window.worker.finished.connect(loop.quit)
        window.worker.failed.connect(loop.quit)
        timer.start(60000)
        loop.exec()
        assert not timer.isActive() or window.result_text.toPlainText()

        crs_path = tmp_path / "botao_CRS.xml"
        fatca_path = tmp_path / "botao_FATCA.xml"
        assert crs_path.exists()
        assert fatca_path.exists()
        assert (tmp_path / "novo_layout_dados_mock_relatorio_auditoria.xlsx").exists()
        assert (tmp_path / "novo_layout_dados_mock_relatorio_auditoria.csv").exists()
        manifest_path = tmp_path / "novo_layout_dados_mock_manifesto_auditoria.json"
        assert manifest_path.exists()

        crs_text = crs_path.read_text(encoding="utf-8")
        fatca_text = fatca_path.read_text(encoding="utf-8")
        assert "AUTO-BR" not in crs_text
        assert "AUTO-BR" not in fatca_text
        identifiers = re.findall(r"<(?:[^:<>]+:)?(?:MessageRefId|DocRefId)>([^<]+)</", crs_text + fatca_text)
        assert len(identifiers) == len(set(identifiers))

        crs = etree.parse(str(crs_path))
        fatca = etree.parse(str(fatca_path))
        account_reports = crs.findall(f".//{{{CRS_NS}}}AccountReport")
        organisations = crs.findall(f".//{{{CRS_NS}}}AccountHolder/{{{CRS_NS}}}Organisation")
        fatca_account_reports = fatca.xpath(".//*[local-name()='AccountReport']")
        assert len(account_reports) == 29
        assert len(organisations) == 11
        assert len(fatca_account_reports) == 0
        assert sum(1 for org in organisations for item in org.findall(f"{{{CRS_NS}}}TIN") if item.get("issuedBy") == "BR") == 11
        assert sum(1 for org in organisations for item in org.findall(f"{{{CRS_NS}}}IN") if item.get("issuedBy") == "BR") == 0
        assert crs.findtext(f".//{{{CRS_NS}}}ReportingFI/{{{CRS_NS}}}IN") == "FI107442"
        assert fatca.find(".//{urn:oecd:ties:fatca:v2}AccountHolder//{urn:oecd:ties:stffatcatypes:v2}TIN") is None

        manifest = json.loads(manifest_path.read_text(encoding="utf-8-sig"))
        assert manifest["identifier_profile"] == "DITC_SEQUENCE"
        assert manifest["final_status"] == "NAO APTO PARA ENVIO"
        assert manifest["counts"]["total de registros recebidos"] == 29
    finally:
        window.close()
        app.processEvents()


def test_botao_simples_respeita_checkboxes_de_crs_e_fatca(tmp_path: Path) -> None:
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    qtwidgets = pytest.importorskip("PySide6.QtWidgets")
    from crs_fatca_generator.gui.main_window import MainWindow

    app = qtwidgets.QApplication.instance() or qtwidgets.QApplication([])
    window = MainWindow()
    captured: dict[str, object] = {}
    try:
        window.excel_path = tmp_path / "entrada.xlsx"
        window.headers = ["DocumentoCliente", "Tipo de documento", "NumConta", "NomeCliente", "SaldoTotal", "Endereco", "Cidade", "Pais"]
        window.sheet_combo.clear()
        window.sheet_combo.addItem("Dados")
        window.crs_check.setChecked(False)
        window.fatca_check.setChecked(True)
        window.configure_simple_outputs = lambda: None  # type: ignore[method-assign]

        def fake_generate(ignore_invalid_records: bool = False) -> None:
            captured["kinds"] = window.kinds()
            captured["declaration"] = window._table_to_profile().declaration

        window.generate_xml = fake_generate  # type: ignore[method-assign]
        window.execute_simple()

        assert captured["kinds"] == ["fatca"]
        assert captured["declaration"] == "fatca"
        assert window.crs_check.isChecked() is False
        assert window.fatca_check.isChecked() is True
    finally:
        window.close()
        app.processEvents()


def test_geracao_reporta_progresso_por_xml_e_conta(tmp_path: Path) -> None:
    profile = infer_default_profile(["DocumentoCliente", "Tipo de documento", "NumConta", "NomeCliente", "SaldoTotal", "Endereco", "Cidade", "Estado", "Pais"])
    profile.output.crs_path = str(tmp_path / "progresso_CRS.xml")
    profile.output.fatca_path = str(tmp_path / "progresso_FATCA.xml")
    rows = [
        {
            "DocumentoCliente": "06360698501",
            "Tipo de documento": "PF",
            "NumConta": "1001",
            "NomeCliente": "Cliente Um",
            "SaldoTotal": "10",
            "Endereco": "Rua A",
            "Cidade": "Sao Paulo",
            "Estado": "SP",
            "Pais": "BR",
            "_excel_row": 2,
        },
        {
            "DocumentoCliente": "09030562595",
            "Tipo de documento": "PF",
            "NumConta": "1002",
            "NomeCliente": "Cliente Dois",
            "SaldoTotal": "20",
            "Endereco": "Rua B",
            "Cidade": "Rio de Janeiro",
            "Estado": "RJ",
            "Pais": "BR",
            "_excel_row": 3,
        },
    ]
    events: list[dict[str, object]] = []

    results = GenerationService(default_crs_schema(), default_fatca_schema()).generate(
        ["crs", "fatca"],
        rows,
        profile,
        Path("entrada.xlsx"),
        overwrite=True,
        progress_callback=events.append,
    )

    assert [result.valid for result in results] == [True, True]
    assert any(event["phase"] == "Escrevendo XML" and event["kind"] == "CRS" and event["processed"] == 1 for event in events)
    assert any(event["phase"] == "Escrevendo XML" and event["kind"] == "FATCA" and event["processed"] == 1 for event in events)
    assert any(event["phase"] == "Montando dados" and event["kind"] == "CRS" and "1001" in str(event["current_record"]) for event in events)
    assert any(event["phase"] == "Validando XSD" and event["kind"] == "FATCA" for event in events)
    assert any(str(event["phase"]).startswith("Preparando dados: documentos") and event["processed"] == 1 for event in events)
    assert any(str(event["phase"]).startswith("Gerando auditoria: Entrada") for event in events)


def test_preparacao_nao_cria_alerta_repetido_por_linha_quando_coluna_ausente() -> None:
    profile = infer_default_profile(["DocumentoCliente", "Tipo de documento", "NumConta", "SaldoTotal", "Pais"])
    rows = [
        {"DocumentoCliente": "06360698501", "Tipo de documento": "PF", "NumConta": f"ACC-{index}", "SaldoTotal": "10", "Pais": "BR", "_excel_row": index}
        for index in range(2, 102)
    ]
    events: list[dict[str, object]] = []

    prepared = DataPreparationService().prepare(rows, profile, progress_callback=events.append)

    rule_01_missing = [
        event
        for event in prepared.events
        if event.rule == "REGRA_01" and event.result == "NAO_AVALIADA_DADO_AUSENTE"
    ]
    rule_03_missing = [
        event
        for event in prepared.events
        if event.rule == "REGRA_03" and event.result == "NAO_AVALIADA_DADO_AUSENTE"
    ]
    assert len(rule_01_missing) == 1
    assert len(rule_03_missing) == 1
    assert any(event["phase"] == "Preparando dados: regras de encerramento CI" and event["processed"] == len(rows) for event in events)


def test_identificador_sequencial_nao_reinicia_busca_a_cada_conta() -> None:
    class MemoryStore:
        def __init__(self) -> None:
            self.values: set[tuple[str, str]] = set()
            self.checks: list[tuple[str, str]] = []

        def exists(self, kind: str, value: str) -> bool:
            self.checks.append((kind, value))
            return (kind, value) in self.values

        def add(self, kind: str, value: str, file_hash: str = "") -> None:
            self.values.add((kind, value))

    store = MemoryStore()
    profile = infer_default_profile(["DocumentoCliente", "Tipo de documento", "NumConta", "SaldoTotal", "Pais"])
    service = MappingService(store)  # type: ignore[arg-type]

    first = service._new_id("crs-account-doc", profile, "")  # noqa: SLF001
    store.add("crs-account-doc", first)
    checks_after_first = len(store.checks)
    second = service._new_id("crs-account-doc", profile, "")  # noqa: SLF001

    assert first != second
    assert len(store.checks) == checks_after_first + 1
    assert first not in [value for _, value in store.checks[checks_after_first:]]


def test_auditoria_grande_usa_xlsx_resumido_e_manifesto_amostral(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(preparation_module, "AUDIT_FULL_XLSX_ROW_LIMIT", 2)
    profile = infer_default_profile(["DocumentoCliente", "Tipo de documento", "NumConta", "SaldoTotal", "Pais"])
    rows = [
        {"DocumentoCliente": "06360698501", "Tipo de documento": "PF", "NumConta": f"ACC-{index}", "SaldoTotal": "10", "Pais": "BR", "_excel_row": index}
        for index in range(2, 6)
    ]
    prepared = DataPreparationService().prepare(rows, profile)
    service = DataPreparationService()

    csv_path, xlsx_path, json_path = service.write_audit(prepared, tmp_path, "grande", profile=profile)

    assert csv_path.exists()
    assert xlsx_path.exists()
    manifest = json.loads(json_path.read_text(encoding="utf-8"))
    assert manifest["audit_xlsx_mode"] == "resumido"
    assert manifest["manifest_detail_mode"] == "amostra"
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        assert "Modo_XLSX" in workbook.sheetnames
        assert "Entrada_Amostra" in workbook.sheetnames
        assert "Entrada" not in workbook.sheetnames
    finally:
        workbook.close()


def test_selecao_do_excel_nao_le_planilha_inteira_na_interface() -> None:
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    qtwidgets = pytest.importorskip("PySide6.QtWidgets")
    from crs_fatca_generator.gui.main_window import MainWindow

    excel_path = Path("ExemplosDados/novo_layout_dados_mock.xlsx")
    assert excel_path.exists()
    app = qtwidgets.QApplication.instance() or qtwidgets.QApplication([])
    window = MainWindow()
    try:
        def fail_if_full_read(*args: object, **kwargs: object) -> list[dict[str, object]]:
            raise AssertionError("A selecao do Excel nao deve chamar read_rows completo na thread da interface.")

        window.excel_reader.read_rows = fail_if_full_read  # type: ignore[method-assign]
        window.excel_path = excel_path
        window.excel_path_edit.setText(str(excel_path))
        window.sheet_combo.clear()
        window.sheet_combo.addItems(window.excel_reader.list_sheets(excel_path))
        window.load_preview()

        assert window.rows == []
        assert 0 < len(window.preview_rows) <= 25
        assert "Previa:" in window.preview_info.text()
    finally:
        window.close()
        app.processEvents()


def test_controlling_person_service_detecta_blocos_dinamicos_e_normaliza() -> None:
    headers = controlling_headers(5)
    values = [""] * len(headers)
    values[0] = "PJ"
    values[1] = "00022003000126"
    values[2] = "ABC HOLDINGS LTDA"
    values[3] = "ACC-CP"
    starts = [index for index, header in enumerate(headers) if index >= 23 and header == "Name Type"]
    first = starts[0]
    values[first + 1 : first + 11] = ["JOAO", "SILVA", "CRS801 - CP of legal person - ownership", "BR", 1234567890, "BR", "OECD304", "BR", "Sao Paulo", "1980-01-02"]
    third = starts[2]
    values[third + 1 : third + 11] = ["MARIA", "SOUZA", "CRS802", "BR", "09030562595", "BR", "OECD304", "BR", "Rio", ""]
    fourth = starts[3]
    values[fourth + 1 : fourth + 11] = ["CARLOS", "LIMA", "CRS803", "BR", "16011329721", "BR", "OECD304", "BR", "Curitiba", "1991-05-06"]
    fifth = starts[4]
    values[fifth + 1 : fifth + 11] = ["ANA", "COSTA", "CRS804", "BR", "52998224725", "BR", "OECD304", "BR", "Salvador", "1975-11-12"]

    assert len(detect_controlling_person_blocks(headers)) == 5
    records, issues, metrics = extract_controlling_persons(row_from_raw(headers, values))

    assert issues == []
    assert [record.result for record in records] == ["INCLUIDO"] * 4
    assert records[0].normalized_tin == "01234567890"
    assert records[0].normalized is True
    assert records[1].controlling_person_type == "CRS802"
    assert records[1].birth_date == ""
    assert metrics["empty_blocks"] == 1
    assert metrics["normalized_cpfs"] == 1


def test_controlling_person_service_valida_cpf_duplicado_e_empresas_diferentes() -> None:
    headers = controlling_headers(2)
    starts = [index for index, header in enumerate(headers) if index >= 23 and header == "Name Type"]
    values = [""] * len(headers)
    values[0] = "PJ"
    values[1] = "00022003000126"
    values[2] = "ABC HOLDINGS LTDA"
    values[3] = "ACC-CP"
    for start in starts:
        values[start + 1 : start + 11] = ["JOAO", "SILVA", "CRS801", "BR", "11111111111", "BR", "OECD304", "BR", "Sao Paulo", ""]

    records, issues, _ = extract_controlling_persons(row_from_raw(headers, values))
    assert len(records) == 2
    assert any("CPF invalido" in issue.message for issue in issues)
    assert any("duplicado" in issue.message for issue in issues)

    values_ok = list(values)
    for start in starts:
        values_ok[start + 5] = "06360698501"
    row_a = row_from_raw(headers, values_ok)
    row_b = row_from_raw(headers, values_ok)
    row_b["AccountNumber"] = "ACC-CP-2"
    row_b["_excel_row"] = 3
    assert extract_controlling_persons(row_a)[1] != []
    values_one = list(values_ok)
    for index in range(starts[1], starts[1] + 11):
        values_one[index] = ""
    assert extract_controlling_persons(row_from_raw(headers, values_one))[1] == []
    assert extract_controlling_persons(row_from_raw(headers, values_one))[0][0].normalized_tin == "06360698501"


def test_fluxo_botao_com_exemplo_novo_layout_gera_dois_controladores_mesmo_account_report(tmp_path: Path) -> None:
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    qtwidgets = pytest.importorskip("PySide6.QtWidgets")
    qtcore = pytest.importorskip("PySide6.QtCore")
    from crs_fatca_generator.gui.main_window import MainWindow

    source = Path("ExemplosDados/exemplo_novo_layout.xlsx")
    excel_path = tmp_path / "exemplo_novo_layout_controladores.xlsx"
    shutil.copy2(source, excel_path)
    workbook = openpyxl.load_workbook(excel_path)
    try:
        sheet = workbook["Planilha1"]
        target_row = 3
        sheet.cell(target_row, 5).value = "true"
        sheet.cell(target_row, 13).value = "CRS101"
        sheet.cell(target_row, 25).value = "JOAO"
        sheet.cell(target_row, 26).value = "SILVA"
        sheet.cell(target_row, 27).value = "CRS801 - CP of legal person - ownership"
        sheet.cell(target_row, 28).value = "BR"
        sheet.cell(target_row, 29).value = 1234567890
        sheet.cell(target_row, 30).value = "BR"
        sheet.cell(target_row, 31).value = "OECD304"
        sheet.cell(target_row, 32).value = "BR"
        sheet.cell(target_row, 33).value = "Sao Paulo"
        sheet.cell(target_row, 34).value = "1980-01-02"
        sheet.cell(target_row, 36).value = "MARIA"
        sheet.cell(target_row, 37).value = "SOUZA"
        sheet.cell(target_row, 38).value = "CRS801"
        sheet.cell(target_row, 39).value = "BR"
        sheet.cell(target_row, 40).value = "09030562595"
        sheet.cell(target_row, 41).value = "BR"
        sheet.cell(target_row, 42).value = "OECD304"
        sheet.cell(target_row, 43).value = "BR"
        sheet.cell(target_row, 44).value = "Rio de Janeiro"
        sheet.cell(target_row, 45).value = "1979-03-04"
        workbook.save(excel_path)
    finally:
        workbook.close()

    app = qtwidgets.QApplication.instance() or qtwidgets.QApplication([])
    window = MainWindow()
    try:
        window.excel_path = excel_path
        window.excel_path_edit.setText(str(excel_path))
        window.sheet_combo.clear()
        window.sheet_combo.addItems(window.excel_reader.list_sheets(excel_path))
        window.load_preview()

        def configure_test_outputs() -> None:
            window.crs_output_edit.setText(str(tmp_path / "controladores_CRS.xml"))
            window.fatca_output_edit.setText(str(tmp_path / "controladores_FATCA.xml"))

        window.configure_simple_outputs = configure_test_outputs  # type: ignore[method-assign]
        button = window.findChild(qtwidgets.QPushButton, "execute_now_button")
        assert button is not None
        button.click()

        loop = qtcore.QEventLoop()
        timer = qtcore.QTimer()
        timer.setSingleShot(True)
        timer.timeout.connect(loop.quit)
        assert window.worker is not None
        window.worker.finished.connect(loop.quit)
        window.worker.failed.connect(loop.quit)
        timer.start(60000)
        loop.exec()

        crs_path = tmp_path / "controladores_CRS.xml"
        fatca_path = tmp_path / "controladores_FATCA.xml"
        assert crs_path.exists()
        assert fatca_path.exists()

        crs = etree.parse(str(crs_path))
        account = crs.xpath(
            ".//*[local-name()='AccountReport'][*[local-name()='AccountNumber' and text()='9834206']]"
        )[0]
        controllers = account.xpath("./*[local-name()='ControllingPerson']")
        assert len(controllers) == 2
        assert account.xpath("count(.//*[local-name()='AccountHolder']/*[local-name()='AcctHolderType' and text()='CRS101'])") == 1
        assert [item.text for item in account.xpath("./*[local-name()='ControllingPerson']/*[local-name()='Individual']/*[local-name()='TIN']")] == [
            "01234567890",
            "09030562595",
        ]
        assert all(item.get("issuedBy") == "BR" for item in account.xpath("./*[local-name()='ControllingPerson']/*[local-name()='Individual']/*[local-name()='TIN']"))

        manifest = json.loads((tmp_path / "exemplo_novo_layout_controladores_manifesto_auditoria.json").read_text(encoding="utf-8-sig"))
        assert manifest["counts"]["empresas com controlling person"] == 1
        assert manifest["counts"]["total de controlling persons recebidos"] == 2
        assert manifest["counts"]["total de controlling persons incluidos"] == 2
        assert manifest["counts"]["maximo de controladores por empresa"] == 2
        assert manifest["controlling_person_reconciliation"][-1]["resultado"] == "ok"
        workbook = openpyxl.load_workbook(tmp_path / "exemplo_novo_layout_controladores_relatorio_auditoria.xlsx", read_only=True)
        try:
            assert "ControllingPersons" in workbook.sheetnames
            assert "Conciliacao_Controladores" in workbook.sheetnames
        finally:
            workbook.close()
    finally:
        window.close()
        app.processEvents()


def test_brl_e_rejeitado_quando_usd_e_obrigatorio(tmp_path: Path) -> None:
    profile = build_sample_profile(tmp_path)
    profile.field_mappings["account.currency"] = MappingRule("fixed", fixed_value="BRL")
    results = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["crs"], [{"_excel_row": 2}], profile, overwrite=True)
    assert results[0].valid is False
    assert any(issue.code == "CUR001" for issue in results[0].issues)


def test_regras_preparacao_removem_duplicadas_e_zeram_saldo() -> None:
    profile = infer_default_profile(["DocumentoCliente", "Tipo de documento", "NumConta", "SaldoTotal", "Pais"])
    rows = [
        {"DocumentoCliente": "063.606.985-01", "Tipo de documento": "PF", "NumConta": 3, "SaldoTotal": "10", "Pais": "BR", "DataHoraEncerramento CI": "2024-12-31", "_excel_row": 2},
        {"DocumentoCliente": "09030562595", "Tipo de documento": "PF", "NumConta": 8, "SaldoTotal": "20", "Pais": "BR", "DataHoraEncerramento CI": "NULL", "Status em 31/12 em CC": "Encerrada", "Encerramento CC": "2025-03-01", "_excel_row": 3},
        {"DocumentoCliente": "16011329721", "Tipo de documento": "PF", "NumConta": 2, "SaldoTotal": "-5.40", "Pais": "BR", "DataHoraEncerramento CI": "2026-01-01", "_excel_row": 4},
        {"DocumentoCliente": "16011329721", "Tipo de documento": "PF", "NumConta": 1, "SaldoTotal": "9", "Pais": "BR", "DataHoraEncerramento CI": "2026-01-01", "_excel_row": 5},
    ]
    prepared = DataPreparationService().prepare(rows, profile)
    assert prepared.issues == []
    assert [row["NumConta"] for row in prepared.rows] == [2]
    assert prepared.rows[0]["SaldoTotal"] == "0.00"
    assert {event.rule for event in prepared.events} >= {"REGRA_01", "REGRA_02", "REGRA_03", "SALDO_NEGATIVO"}


def test_cpf_invalido_bloqueia_geracao(tmp_path: Path) -> None:
    profile = infer_default_profile(["DocumentoCliente", "Tipo de documento", "NumConta", "NomeCliente", "SaldoTotal", "Endereco", "Cidade", "Estado", "Pais"])
    profile.output.crs_path = str(tmp_path / "crs.xml")
    rows = [
        {
            "DocumentoCliente": "11111111111",
            "Tipo de documento": "PF",
            "NumConta": "1",
            "NomeCliente": "Cliente Teste",
            "SaldoTotal": "1",
            "Endereco": "Rua A",
            "Cidade": "Sao Paulo",
            "Estado": "SP",
            "Pais": "BR",
            "_excel_row": 2,
        }
    ]
    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["crs"], rows, profile, Path("entrada.xlsx"), overwrite=True)[0]
    assert result.valid is False
    assert any(issue.code == "DOC001" for issue in result.issues)
    assert not (tmp_path / "crs.xml").exists()


def test_geracao_pode_ignorar_linha_com_controlador_invalido(tmp_path: Path) -> None:
    headers = controlling_headers(1)
    bad_values = [""] * len(headers)
    bad_values[0] = "PJ"
    bad_values[1] = "00022003000126"
    bad_values[2] = "EMPRESA COM CONTROLADOR INVALIDO"
    bad_values[3] = "ACC-BAD"
    bad_values[4] = "true"
    bad_values[8] = "BR"
    bad_values[12] = "CRS101"
    bad_values[20] = "10"
    bad_values[21] = "USD"
    bad_values[22] = "OECD1"
    bad_values[23:34] = ["OECD202", "JOAO", "SILVA", "CRS801", "BR", "11111111111", "BR", "OECD304", "BR", "Sao Paulo", "1980-01-01"]
    bad_row = row_from_raw(headers, bad_values)

    good_values = [""] * len(headers)
    good_values[0] = "PF"
    good_values[1] = "06360698501"
    good_values[2] = "CLIENTE OK"
    good_values[3] = "ACC-OK"
    good_values[4] = "true"
    good_values[8] = "BR"
    good_values[20] = "25"
    good_values[21] = "USD"
    good_values[22] = "OECD1"
    good_row = row_from_raw(headers, good_values)
    good_row["_excel_row"] = 3

    profile = infer_default_profile(headers)
    profile.output.crs_path = str(tmp_path / "ignorar_CRS.xml")
    profile.output.fatca_path = str(tmp_path / "ignorar_FATCA.xml")
    service = GenerationService(default_crs_schema(), default_fatca_schema())

    blocked = service.generate(["crs"], [bad_row, good_row], profile, Path("entrada.xlsx"), overwrite=True)
    assert blocked[0].valid is False
    assert any(issue.code == "CP001" and issue.excel_row == 2 for issue in blocked[0].issues)
    assert not (tmp_path / "ignorar_CRS.xml").exists()

    generated = service.generate(
        ["crs"],
        [bad_row, good_row],
        profile,
        Path("entrada.xlsx"),
        overwrite=True,
        ignore_invalid_records=True,
    )
    assert generated[0].valid is True
    crs = etree.parse(str(tmp_path / "ignorar_CRS.xml"))
    assert crs.xpath("count(.//*[local-name()='AccountNumber' and text()='ACC-OK'])") == 1
    assert crs.xpath("count(.//*[local-name()='AccountNumber' and text()='ACC-BAD'])") == 0
    manifest = json.loads((tmp_path / "entrada_manifesto_auditoria.json").read_text(encoding="utf-8"))
    assert manifest["counts"]["total excluido"] == 1
    assert any(item["regra"] == "ERRO_IGNORADO" for item in manifest["excluded_records"])


def _fatca_profile(tmp_path: Path) -> object:
    headers = ["DocumentoCliente", "Tipo de documento", "NumConta", "NomeCliente", "SaldoTotal", "Endereco", "Cidade", "Estado", "Pais"]
    profile = infer_default_profile(headers)
    profile.output.fatca_path = str(tmp_path / "fatca.xml")
    profile.output.crs_path = str(tmp_path / "crs.xml")
    return profile


def _fatca_row(document: str = "06360698501", person_type: str = "PF") -> dict[str, object]:
    return {
        "DocumentoCliente": document,
        "Tipo de documento": person_type,
        "NumConta": "11917679",
        "NomeCliente": "Cliente Teste",
        "SaldoTotal": "1",
        "Endereco": "Rua A",
        "Cidade": "Sao Paulo",
        "Estado": "SP",
        "Pais": "BR",
        "_excel_row": 2,
    }


def test_fatca_nao_usa_cpf_como_us_tin(tmp_path: Path) -> None:
    profile = _fatca_profile(tmp_path)
    profile.field_mappings["fatca.us_tin"] = MappingRule("fixed", fixed_value="06360698501")
    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["fatca"], [_fatca_row()], profile, Path("entrada.xlsx"), overwrite=True)[0]
    assert result.valid is False
    assert any(issue.code == "FATCA_TIN001" for issue in result.issues)


def test_fatca_nao_usa_cnpj_como_us_tin(tmp_path: Path) -> None:
    profile = _fatca_profile(tmp_path)
    profile.field_mappings["fatca.us_tin"] = MappingRule("fixed", fixed_value="07859988000198")
    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["fatca"], [_fatca_row("07859988000198", "PJ")], profile, Path("entrada.xlsx"), overwrite=True)[0]
    assert result.valid is False
    assert any(issue.code == "FATCA_TIN001" for issue in result.issues)


def test_fatca_us_tin_valido_preserva_texto_e_zeros(tmp_path: Path) -> None:
    profile = _fatca_profile(tmp_path)
    profile.field_mappings["fatca.us_tin"] = MappingRule("fixed", fixed_value="012-34-6789")
    profile.field_mappings["fatca.us_tin_status"] = MappingRule("fixed", fixed_value="INFORMED")
    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["fatca"], [_fatca_row()], profile, Path("entrada.xlsx"), overwrite=True)[0]
    assert result.valid is True
    tree = etree.parse(str(tmp_path / "fatca.xml"))
    holder_tin = tree.findtext(".//{urn:oecd:ties:fatca:v2}AccountHolder//{urn:oecd:ties:stffatcatypes:v2}TIN")
    assert holder_tin == "012-34-6789"


def test_fatca_us_tin_ausente_omite_tin_e_registra_pendencia(tmp_path: Path) -> None:
    profile = _fatca_profile(tmp_path)
    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["fatca"], [_fatca_row()], profile, Path("entrada.xlsx"), overwrite=True)[0]
    assert result.valid is True
    tree = etree.parse(str(tmp_path / "fatca.xml"))
    assert tree.find(".//{urn:oecd:ties:fatca:v2}AccountHolder//{urn:oecd:ties:stffatcatypes:v2}TIN") is None
    xml_text = (tmp_path / "fatca.xml").read_text(encoding="utf-8")
    assert "<sfa:TIN>NULL</sfa:TIN>" not in xml_text
    assert "<sfa:TIN>N/A</sfa:TIN>" not in xml_text
    assert ">000000000<" not in xml_text
    workbook = openpyxl.load_workbook(tmp_path / "entrada_relatorio_auditoria.xlsx")
    try:
        assert "Pendencias US Tax ID" in workbook.sheetnames
        assert workbook["Pendencias US Tax ID"].max_row == 2
    finally:
        workbook.close()


def test_fatca_us_tin_ausente_com_bloqueio_impede_geracao(tmp_path: Path) -> None:
    profile = _fatca_profile(tmp_path)
    profile.field_mappings["fatca.missing_us_tin_policy"] = MappingRule("fixed", fixed_value="BLOCK_GENERATION")
    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["fatca"], [_fatca_row()], profile, Path("entrada.xlsx"), overwrite=True)[0]
    assert result.valid is False
    assert any(issue.code == "FATCA_TIN001" for issue in result.issues)
    assert not (tmp_path / "fatca.xml").exists()


def test_fatca_planilha_antiga_nao_copia_documento_para_us_tax_id(tmp_path: Path) -> None:
    profile = _fatca_profile(tmp_path)
    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["fatca"], [_fatca_row()], profile, Path("entrada.xlsx"), overwrite=True)[0]
    tree = etree.parse(str(tmp_path / "fatca.xml"))
    assert result.valid is True
    assert tree.findtext(".//{urn:oecd:ties:fatca:v2}AccountHolder//{urn:oecd:ties:stffatcatypes:v2}TIN") is None


def test_auditoria_completa_e_manifesto_json(tmp_path: Path) -> None:
    profile = _fatca_profile(tmp_path)
    rows = [_fatca_row()]
    result = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["crs", "fatca"], rows, profile, Path("entrada.xlsx"), overwrite=True)
    assert [item.valid for item in result] == [True, True]
    workbook = openpyxl.load_workbook(tmp_path / "entrada_relatorio_auditoria.xlsx", read_only=True)
    try:
        expected = {"Resumo", "Entrada", "Decisoes", "Exclusoes", "Transformacoes", "CRS", "FATCA", "Identificadores", "Validacao_XSD", "Conciliacao", "Pendencias"}
        assert expected.issubset(set(workbook.sheetnames))
    finally:
        workbook.close()
    manifest = (tmp_path / "entrada_manifesto_auditoria.json").read_text(encoding="utf-8")
    assert '"processing_id"' in manifest
    assert '"xsd_validation"' in manifest


def test_xml_invalido_e_reportado_pelo_xsd(tmp_path: Path) -> None:
    invalid = tmp_path / "invalid_crs.xml"
    invalid.write_text("<CRS_OECD/>", encoding="utf-8")
    issues = XmlValidator().validate_file(invalid, default_crs_schema(), "crs")
    assert issues
    assert issues[0].code == "XSD"


def test_crs_xsd_embarcado_permite_tin_para_titular_pj_sem_remover_in() -> None:
    schema_text = default_crs_schema().read_text(encoding="utf-8")
    start = schema_text.index('name="OrganisationParty_Type"')
    block = schema_text[start : schema_text.index('name="CorrectableOrganisationParty_Type"', start)]
    assert 'name="IN"' in block
    assert 'name="TIN"' in block


def test_fluxo_simples_gera_ao_lado_do_excel(tmp_path: Path) -> None:
    excel_path = tmp_path / "novo_layout_dados_mock.xlsx"
    shutil.copy2(Path("ExemplosDados/novo_layout_dados_mock.xlsx"), excel_path)
    assert generate_from_excel(excel_path) == 0
    assert (excel_path.parent / "XML_Gerados" / f"{excel_path.stem}_CRS.xml").exists()
    assert (excel_path.parent / "XML_Gerados" / f"{excel_path.stem}_FATCA.xml").exists()


def test_configuracao_portatil_define_saida_sem_admin(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    config_path = tmp_path / "configuracao_portatil.json"
    output_root = tmp_path / "saida_usuario"
    config_path.write_text(
        json.dumps(
            {
                "preferir_saida_ao_lado_do_excel": False,
                "pasta_saida_padrao": str(output_root),
                "pasta_dados_local": str(tmp_path / "dados_usuario"),
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setenv("CIINTEGRACAO_CONFIG", str(config_path))
    output = simple_output_paths(tmp_path / "entrada_teste.xlsx")
    assert Path(output.crs_path).parent == output_root / "entrada_teste"
    assert Path(output.fatca_path).parent == output_root / "entrada_teste"


def test_fluxo_zero_config_faz_fallback_automatico_para_pasta_do_usuario(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    excel_path = tmp_path / "entrada_teste.xlsx"
    monkeypatch.setenv("CIINTEGRACAO_CONFIG", str(tmp_path / "config_inexistente.json"))
    monkeypatch.delenv("CIINTEGRACAO_OUTPUT_DIR", raising=False)
    blocked = excel_path.parent / "XML_Gerados"

    def fake_writable(path: Path) -> bool:
        return path != blocked

    monkeypatch.setattr(path_config, "is_writable_dir", fake_writable)
    output = simple_output_paths(excel_path)
    assert Path(output.crs_path).parent.name == excel_path.stem
    assert "CRS_FATCA_XML_Gerados" in str(Path(output.crs_path).parent)
    assert Path(output.fatca_path).parent == Path(output.crs_path).parent


def test_launcher_cmd_forca_dados_no_perfil_do_usuario() -> None:
    text = Path("Abrir_Gerador_CRS_FATCA.cmd").read_text(encoding="utf-8")
    assert "CIINTEGRACAO_DATA_DIR=%LOCALAPPDATA%\\CRS_FATCA_XML_Generator" in text
    assert "/D \"%~dp0\"" in text


def test_colunas_faltantes_sao_reportadas() -> None:
    headers = ["DocumentoCliente", "NumConta", "NomeCliente"]
    missing = missing_simple_columns(headers)
    assert missing == ["Tipo de documento", "SaldoTotal", "Endereco", "Cidade", "Pais"]


def test_doc_ref_duplicado_e_reportado(tmp_path: Path) -> None:
    profile = build_sample_profile(tmp_path)
    profile.field_mappings["account.doc_ref_id"] = MappingRule("fixed", fixed_value="DOC-DUP")
    rows = [{"Account number*": "A1", "_excel_row": 2}, {"Account number*": "A2", "_excel_row": 3}]
    results = GenerationService(default_crs_schema(), default_fatca_schema()).generate(["crs"], rows, profile, overwrite=True)
    assert results[0].valid is False
    assert any(issue.code == "DOC001" for issue in results[0].issues)


def test_schema_enum_extraido_do_xsd() -> None:
    enums = SchemaInspector().enums(default_crs_schema())
    assert "CRS701" in enums["CrsMessageTypeIndic_EnumType"]
    assert "OECD606" in enums["AcctNumberType_EnumType"]


def test_xsd_ausente_falha_com_erro_claro(tmp_path: Path) -> None:
    with pytest.raises(OSError):
        SchemaInspector().enums(tmp_path / "ausente.xsd")
